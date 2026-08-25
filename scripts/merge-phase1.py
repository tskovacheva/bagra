#!/usr/bin/env python3
"""Merge the phase 1 workbook into seed/plants.json.

FILLS ONLY. A cell that would change something already recorded is held back
and printed, never applied. The rule is the one the whole project runs on: the
workbook is a draft to be checked against the model, not a decision. Thirteen
cells in this return would have overwritten a recorded value, and one of them
contradicts another field in its own record.

Idempotent: run it twice and the second run reports nothing to do.
"""
import json, re, sys, collections
from openpyxl import load_workbook

BOOK = '/mnt/user-data/uploads/Багра-липсващи-данни-фаза-1-попълнени.xlsx'
PACK = 'seed/plants.json'

wb = load_workbook(BOOK)
pack = json.load(open(PACK))
plants = {p['code']: p for p in pack['plants']}

# The vocabulary, read from the source rather than retyped — a merge script with
# its own copy of the codes is a merge script that drifts from the application.
vocab = collections.defaultdict(set)
for m in re.finditer(r"V\('([a-z_]+)',\s*'([a-zA-Z0-9_]+)',", open('vocab.js').read()):
    vocab[m.group(1)].add(m.group(2))

filled, held, skipped = [], [], 0


def norm_range(text):
    """`80–80` is not a range, it is 80.

    A degenerate span reads on screen as „80–80 °C", which looks like a range
    somebody measured twice rather than a single figure. Written as one number
    it says what it means.
    """
    if text in (None, ''):
        return None
    t = str(text).replace('-', '–').replace(' ', '')
    if '–' in t:
        lo, hi = t.split('–', 1)
        lo, hi = int(lo), int(hi)
        return {'min': lo} if lo == hi else {'min': lo, 'max': hi}
    return {'min': int(t)}


def flat(span):
    """`{min:80, max:80}` and `{min:80}` are the same figure written two ways.

    Three of these were already in the pack and one arrived in the workbook.
    Comparing them literally would report a disagreement where there is none,
    and a merge that cries wolf gets its warnings skimmed.
    """
    if not span:
        return None
    if span.get('max') in (None, span.get('min')):
        return {'min': span['min']}
    return {'min': span['min'], 'max': span['max']}


def same_span(current, incoming):
    if current is None or incoming is None:
        return current == incoming
    return flat(current) == flat(incoming)


def offer(where, field, current, incoming, apply):
    """Fill an empty field; hold anything that would change a recorded value."""
    global skipped
    if incoming in (None, '', []):
        return
    if current in (None, '', []):
        apply()
        filled.append((where, field, incoming))
        return
    if isinstance(current, dict) or isinstance(incoming, dict):
        if same_span(current, incoming):
            skipped += 1
            return
    elif str(current).strip() == str(incoming).strip():
        skipped += 1
        return
    held.append((where, field, current, incoming))


# ---------------------------------------------------------------- sheet 4
ws = wb['4 Поле по растение']
for r in range(7, ws.max_row + 1):
    v = [ws.cell(row=r, column=c).value for c in range(1, 11)]
    if not v[1] or str(v[0]).startswith('ПРИМЕР'):
        continue
    p = plants.get(v[1])
    if not p:
        held.append((v[1], 'plant', 'not in the pack', '—'))
        continue
    w = v[1]

    offer(w, 'family', p.get('family'), v[2],
          lambda p=p, x=v[2]: p.__setitem__('family', x))
    offer(w, 'dyeClass', p.get('dyeClass'), v[3],
          lambda p=p, x=v[3]: p.__setitem__('dyeClass', x))

    roles = [x.strip() for x in str(v[4] or '').split(',') if x.strip()]
    offer(w, 'compositionalRole', p.get('compositionalRole') or None, roles or None,
          lambda p=p, x=roles: p.__setitem__('compositionalRole', x))

    offer(w, 'lightfastness', p.get('lightfastness'), v[5],
          lambda p=p, x=v[5]: p.__setitem__('lightfastness', x))
    offer(w, 'washfastness', p.get('washfastness'), v[6],
          lambda p=p, x=v[6]: p.__setitem__('washfastness', x))

    tox = p.setdefault('toxicity', {})
    offer(w, 'toxicity.level', tox.get('level'), v[7],
          lambda t=tox, x=v[7]: t.__setitem__('level', x))
    prec = [x.strip() for x in str(v[8] or '').split(',') if x.strip()]
    offer(w, 'toxicity.precautions', tox.get('precautions') or None, prec or None,
          lambda t=tox, x=prec: t.__setitem__('precautions', x))

# ---------------------------------------------------------------- sheet 5
ws = wb['5 Поле по част']
for r in range(7, ws.max_row + 1):
    v = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    if not v[1] or str(v[0]).startswith('ПРИМЕР'):
        continue
    p = plants.get(v[1])
    pt = next((x for x in (p or {}).get('parts', []) if x['partCode'] == v[2]), None)
    if not pt:
        held.append((f'{v[1]}/{v[2]}', 'part', 'not on this plant', '—'))
        continue
    w = f'{v[1]}/{v[2]}'

    chem = []
    for chunk in str(v[3] or '').split(','):
        chunk = chunk.strip()
        if not chunk:
            continue
        cl, lv = (x.strip() for x in chunk.split('/', 1))
        chem.append({'classCode': cl, 'level': lv})
    offer(w, 'chemistry', pt.get('chemistry') or None, chem or None,
          lambda pt=pt, x=chem: pt.__setitem__('chemistry', x))

    dosing = pt.get('dosing') or []
    cur = dosing[0] if dosing else {}
    if v[4] is not None and not dosing:
        pt['dosing'] = [{'min': v[4], 'max': v[5], 'condition': v[6] or 'dried'}]
        filled.append((w, 'dosing', f'{v[4]}–{v[5]}% WOF, {v[6]}'))
    else:
        offer(w, 'dosing.min', cur.get('min'), v[4], lambda c=cur, x=v[4]: c.__setitem__('min', x))
        offer(w, 'dosing.max', cur.get('max'), v[5], lambda c=cur, x=v[5]: c.__setitem__('max', x))
        offer(w, 'dosing.condition', cur.get('condition'), v[6],
              lambda c=cur, x=v[6]: c.__setitem__('condition', x))

    # A FILL CAN CONTRADICT THE RECORD TOO.
    #
    # „Only fill what is empty" is not enough on its own. Safflower's dyeing
    # temperature was empty, so 70–75 °C arrived as a fill and passed every
    # check — while the same record's `extractionModes` says `cold`, and its
    # colour note says the red comes from an alkaline extraction. Carthamin is
    # drawn out cold; heat destroys it. The number was legal, the code was
    # known, and the record no longer agreed with itself.
    #
    # Caught by the guard added in the same session and repeated here, because
    # a merge that writes a fault and relies on a later layer to notice is a
    # merge that has to be undone by hand.
    cold = (pt.get('extractionModes') or []) == ['cold']
    for field, cell in (('tempExtractC', v[7]), ('tempDyeC', v[8])):
        span = norm_range(cell)
        if cold and span and span['min'] > 40:
            held.append((w, field, 'restricted to COLD extraction', span))
            continue
        offer(w, field, pt.get(field), span,
              lambda pt=pt, f=field, x=span: pt.__setitem__(f, x))

    if cold and v[9] and v[9] > 40:
        held.append((w, 'softMaxTempC', 'restricted to COLD extraction', v[9]))
    else:
        offer(w, 'softMaxTempC', pt.get('softMaxTempC'), v[9],
              lambda pt=pt, x=v[9]: pt.__setitem__('softMaxTempC', x))

# The three degenerate ranges already in the pack, written as what they are.
# Not a change of meaning: on screen „80–80 °C" reads as a range somebody
# measured twice rather than as one figure.
tidied = 0
for p in pack['plants']:
    for pt in p['parts']:
        for f in ('tempExtractC', 'tempDyeC'):
            v = pt.get(f)
            if v and v.get('max') == v.get('min'):
                pt[f] = {'min': v['min']}
                tidied += 1

# ---------------------------------------------------------------- write
if '--apply' in sys.argv:
    pack['packVersion'] = sys.argv[sys.argv.index('--version') + 1] \
        if '--version' in sys.argv else pack['packVersion']
    json.dump(pack, open(PACK, 'w'), ensure_ascii=False, indent=1)

print(f'FILLED   {len(filled)}   (empty → value)')
print(f'ALREADY  {skipped}   (the workbook agrees with the pack)')
print(f'TIDIED   {tidied}   (80–80 written as 80)')
print(f'HELD     {len(held)}   (would change a recorded value — NOT applied)')
print()
by = collections.Counter(f for _, f, _ in filled)
for k, n in by.most_common():
    print(f'  {k:22} {n}')
print()
for w, f, cur, inc in held:
    print(f'  HELD  {w:30} {f:20} {cur!r} → {inc!r}')
if '--apply' not in sys.argv:
    print('\n(dry run — pass --apply to write)')
