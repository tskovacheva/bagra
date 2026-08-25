#!/usr/bin/env python3
"""Merge phase 3 into the combination pack.

Three sources, one script:

  sheet 1  — the fibre and mordant a recorded result was got on (fills only)
  sheet 3  — new combination records for plants that had none
  the docx — the eco print library: 25 plants no workbook could supply, because
             no book has them. They come from bundles.

FILLS ONLY on anything already recorded. A new record is written only where none
exists with that key.
"""
import json, re, sys, collections
from openpyxl import load_workbook

BOOK = '/mnt/user-data/uploads/Багра-фаза-3-попълнено-сигурни.xlsx'

wb = load_workbook(BOOK)
combos_pack = json.load(open('seed/combinations.json'))
sources_pack = json.load(open('seed/sources.json'))
plants = {p['code']: p for p in json.load(open('seed/plants.json'))['plants']}
combos = {r['code']: r for r in combos_pack['combinations']}
known_sources = {s['code'] for s in sources_pack['sources']}

filled, held, created, skipped = [], [], [], 0

CITATIONS = {
    'Jenny Dean': ('jenny-dean-wild-colour', None, None, None),
    'Journal of Natural Fibers — Salvia': ('eser-salvia-2017', None, None, None),
    'Catharine Ellis, Natural Dye': ('ellis-natural-dye', 'book',
        'Natural Dye: Experiments and Results', 'Catharine Ellis'),
    'Luhamaa et al., Heritage 2025': ('luhamaa-frangula-2025', 'reference',
        'Alder buckthorn bark on wool', 'Luhamaa et al., Heritage'),
    'Denver Botanic Gardens': ('denver-dye-garden', 'site',
        'Dye Garden', 'Denver Botanic Gardens'),
    'Cuce, Journal of Natural Fibers': ('cuce-rosa-canina', 'reference',
        'Rosa canina leaf extract on wool', 'Cuce, Journal of Natural Fibers'),
    'Kirby & Saunders, Heritage 2025': ('kirby-brazilwood-2025', 'reference',
        'Brazilwood on alum-mordanted wool', 'Kirby & Saunders, Heritage'),
    'CAMEO': ('cameo-mfa', None, None, None),
}


def cite(text):
    if not text:
        return None
    text = str(text).strip()
    url = None
    m = re.search(r'https?://\S+', text)
    if m:
        url = m.group(0).rstrip('.,;')
    for prefix, (code, kind, name, author) in CITATIONS.items():
        if text.startswith(prefix):
            if kind and code not in known_sources:
                sources_pack['sources'].append({
                    'code': code, 'kind': kind, 'name': name, 'author': author,
                    'url': url or '',
                    'note': {'bg': 'Регистриран при сливането на фаза 3. Цитиран за '
                                   'конкретна двойка влакно и мордант.',
                             'en': 'Registered when phase 3 was merged. Cited for one '
                                   'fibre-and-mordant pairing.'}})
                known_sources.add(code)
            return code
    return None


def offer(where, field, current, incoming, apply):
    global skipped
    if incoming in (None, '', []):
        return
    if current in (None, '', []):
        apply(); filled.append((where, field, incoming)); return
    if str(current).strip() == str(incoming).strip():
        skipped += 1; return
    held.append((where, field, str(current)[:32], str(incoming)[:32]))


# ---------------------------------------------------------------- sheet 1
ws = wb['1 Влакно и мордант']
for r in range(7, ws.max_row + 1):
    v = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    code = v[0]
    if not code or str(code).startswith('ПРИМЕР'):
        continue
    if not any(x not in (None, '') for x in v[6:11]):
        continue
    rec = combos.get(code)
    if not rec:
        held.append((code, 'record', 'not in the pack', '—')); continue
    key = rec['key']
    offer(code, 'fibreClass', key.get('fibreClass'), v[6],
          lambda k=key, x=v[6]: k.__setitem__('fibreClass', x))
    offer(code, 'mordantCode', key.get('mordantCode'), v[7],
          lambda k=key, x=v[7]: k.__setitem__('mordantCode', x))
    offer(code, 'mordantBand', key.get('mordantBand'), v[8],
          lambda k=key, x=v[8]: k.__setitem__('mordantBand', x))
    scode = cite(v[10])
    if scode:
        offer(code, 'learnedFrom', rec.get('learnedFrom'), scode,
              lambda rr=rec, x=scode: rr.__setitem__('learnedFrom', x))
    elif v[10]:
        held.append((code, 'source', 'unrecognised', str(v[10])[:32]))
    if v[9]:
        cur = (rec.get('notes') or {}).get('bg')
        if cur:
            held.append((code, 'notes', cur[:32], str(v[9])[:32]))
        else:
            rec.setdefault('notes', {'bg': '', 'en': ''})['bg'] = str(v[9]).strip()


def make(plant, part, fibre, mordant, band, process, colour, variation,
         confidence, source, medium=None, quality=None, note=''):
    code = '_'.join(x for x in [plant, part, mordant or 'nomordant', process] if x)
    if code in combos:
        held.append((code, 'record', 'a record with this code exists', '—'))
        return None
    rec = {
        'code': code,
        'key': {
            'dyeSource': {'plantId': f'seed:{plant}', 'partCode': part},
            'fibreClass': fibre, 'fibreCode': None,
            'mordantCode': mordant, 'mordantBand': band,
            'processCode': process, 'blanket': None, 'medium': medium,
        },
        'expected': {
            'colourText': {'bg': colour, 'en': ''},
            # No hex. The sources describe colour in WORDS, and turning „наситено
            # златисто жълто" into a hex value would be the application inventing
            # a measurement nobody made (§13ax). Left empty; the swatch does not
            # draw, which is the honest rendering of „nobody has measured this".
            'swatchHex': '',
            'variation': {'bg': variation, 'en': ''},
            'printQuality': quality,
            'lightfastness': '', 'washfastness': '',
        },
        'influences': [], 'confidence': confidence, 'learnedFrom': source,
        'notes': {'bg': note, 'en': ''},
    }
    combos_pack['combinations'].append(rec)
    combos[code] = rec
    created.append(code)
    return rec


# ---------------------------------------------------------------- sheet 3
ws = wb['3 Липсващи растения']
for r in range(7, ws.max_row + 1):
    v = [ws.cell(row=r, column=c).value for c in range(1, 16)]
    if str(v[0] or '').startswith('ПРИМЕР'):
        continue
    if not any(x not in (None, '') for x in v[3:15]):
        continue
    plant = v[1]
    if plant not in plants:
        held.append((str(plant), 'plant', 'not in the pack', '—')); continue
    if v[3] not in [pt['partCode'] for pt in plants[plant]['parts']]:
        held.append((f'{plant}/{v[3]}', 'part', 'not on this plant', '—')); continue
    make(plant, v[3], v[5], v[6], v[7], v[4] or 'immersion',
         str(v[10] or '').strip(), str(v[12] or '').strip(),
         v[13] or 'literature', cite(v[14]) or str(v[14] or '').strip(),
         medium=v[8] or None)


# ---------------------------------------------------------------- eco print
#
# Read out of the document BY HAND rather than parsed. Twenty-five entries is
# few enough to read, and every one needed a decision no regular expression can
# make: which part the source actually means, whether a colour belongs to alum
# or to no mordant at all, and which sentences describe the PRINT rather than
# the colour. A parser would have produced fifty records and no judgement.
ECO = [
    ('anthemis_tinctoria', 'flower', 'sharp',
     'Кръгли ярки петна от кошничките; нарязаните листа оставят ефирен силует.',
     'ярко до слънчево жълто', 'маслиненозелено'),
    ('betula_pendula', 'leaf', 'sharp',
     'Фини линии на жилките. Тънките листа искат силно притискане в рулото.',
     'меко жълто-зелено', 'ръждиво-кафяво или сиво-синкаво'),
    ('achillea_millefolium', 'leaf', 'sharp',
     'Перестите листа дават дантелен силует, наподобяващ папрат.',
     'светложълто до тревистозелено', 'тъмно маслинено или сиво-зелено'),
    ('tanacetum_vulgare', 'leaf', 'sharp',
     'Нарязаните листа дават графичен назъбен отпечатък; цветните бутони правят кръгове.',
     'наситено златисто жълто', 'дълбоко тъмнозелено до кафяво'),
    ('crataegus_monogyna', 'leaf', 'sharp',
     'Малките нарязани листа дават отчетливи контури.',
     'меко бежово, охра, жълтеникаво', 'тъмносиво до черно по жилките'),
    ('salvia_officinalis', 'leaf', 'sharp',
     'Релефните пухкави листа дават ясна мрежа от жилки.',
     'меко жълто-сиво или зелено', 'драматично тъмносиво'),
    ('cornus_mas', 'leaf', 'sharp',
     'Дъговидните жилки се изписват перфектно.',
     'светлозелено или жълтеникаво', 'тъмно кафяво-сиво'),
    ('eucalyptus_spp', 'leaf', 'sharp',
     'Едно от най-мощните растения за еко принт. И двете страни печатат силно, '
     'поради обилните етерични масла.',
     'наситено червено, оранжево, ръждиво и шоколадово кафяво', None),
    ('hypericum_perforatum', 'leaf', 'soft',
     'Малки листа с фини полупрозрачни жлези; цветчетата оставят нежни петна.',
     'червеникаво, бордо или маслинено', 'тъмнозелено'),
    ('geranium_macrorrhizum', 'leaf', 'sharp',
     'Широките длановидни листа правят големи силуети; жилките излизат при стегнато руло.',
     'нежно зелено до кафяво-зелено', 'сиво-черно, графитно'),
    ('rhamnus_cathartica', 'leaf', 'sharp',
     'Силно изразени жилки. Плодовете дават историческото „сап-зелено".',
     'жълто-зелено', None),
    ('castanea_sativa', 'leaf', 'sharp',
     'Големите назъбени листа дават скелетен отпечатък на зъбците и централната жилка.',
     'златисто-кафяво', 'дълбоко мастилено черно'),
    ('urtica_dioica', 'leaf', 'sharp',
     'Подробна назъбена текстура. Парещите власинки омекват под парата.',
     'бутилково или маслинено зелено', 'тъмно мастилено зелено'),
    ('coreopsis_tinctoria', 'flower', 'diffuse',
     'Цветовете се разстилат под натиска и оставят експлозивни ярки петна.',
     'огнено оранжево, теракота, наситено жълто', 'тъмно кафяво-оранжево'),
    ('cosmos_sulphureus', 'flower', 'soft',
     'Цели притиснати цветове дават звездовидни петна.',
     'ярко мандаринено оранжево', 'маслинено-кафяви земни тонове'),
    ('corylus_avellana', 'leaf', 'sharp',
     'Широките мъхести листа имат подчертан релеф и дават чист отпечатък.',
     'меко бежово и кремаво', 'сиво-кафяво'),
    ('tilia_cordata', 'leaf', 'soft',
     'Сърцевидните листа дават фини, нежни отпечатъци.',
     'светло деликатно жълто до кремаво', 'меко сиво по жилките'),
    ('calendula_officinalis', 'flower', 'soft',
     'Разпръснати слънчеви петна или ясен кръгъл отпечатък от цвета.',
     'меко жълто до нежно оранжево', 'маслинен нюанс'),
    ('juglans_regia', 'leaf', 'sharp',
     'Дебелите жилки правят перфектни линии. Зелената черупка на плода също печата силно.',
     'тъмнокафяво и шоколадово', 'почти гарваново черно'),
    ('rosa_spp', 'leaf', 'sharp',
     'Сложните листа с назъбен край правят изящни отпечатъци. Цветовете от тъмни рози '
     'могат да оставят лилави или синкави петна.',
     'кафяво-зелено', 'силно тъмнозелено до черно'),
    ('cotinus_coggygria', 'leaf', 'sharp',
     'Най-детайлните отпечатъци на жилки сред българските растения.',
     'златисто жълто до оранжево', 'тъмносиво, лилаво-черно или графитно'),
    ('rhus_coriaria', 'leaf', 'sharp',
     'Сложните перести листа дават графичен ред от малки листенца.',
     'светлокафяво до сиво-жълто', 'дълбоко сиво-черно'),
    ('tagetes_erecta', 'flower', 'soft',
     'Цветовете правят плътни жълти петна; листата дават ясни тънки контури.',
     'интензивно ярко жълто и оранжево', 'дълбоко маслинено или тъмнозелено'),
    ('cornus_sanguinea', 'leaf', 'sharp',
     'Красиви дъговидни жилки. Летораслите и есенните листа съдържат антоцианини.',
     'розово-кафяво или червеникаво', 'тъмно кафяво-сиво'),
    ('malus_domestica', 'leaf', 'soft',
     'Плътни, леко мъхнати отдолу. Мъхът се почиства, за да излезе чист отпечатък.',
     'меко жълто до охра', 'земно кафяво'),
]

ECO_SOURCE = 'crafty-place-ecoprint'
if ECO_SOURCE not in known_sources:
    sources_pack['sources'].append({
        'code': ECO_SOURCE, 'kind': 'reference',
        'name': 'Еко принт библиотека — преглед на 25 растения',
        'author': 'Crafty Place', 'url': '',
        'note': {'bg': 'Собствен преглед на еко принт поведението на двайсет и пет растения, '
                       'съставен от практика и от публикувани отпечатъци. Описва отпечатък и '
                       'цвят; не назовава влакно, затова записите оттук мълчат по този въпрос.',
                 'en': 'An in-house review of how twenty-five plants behave in eco print, '
                       'assembled from practice and from published prints. It describes the '
                       'print and the colour; it does not name a fibre, so records from it '
                       'stay silent on that.'}})
    known_sources.add(ECO_SOURCE)

eco_made, eco_skipped = 0, []
for plant, part, quality, note, alum, iron in ECO:
    if plant not in plants:
        eco_skipped.append((plant, 'plant not in the pack')); continue
    if part not in [pt['partCode'] for pt in plants[plant]['parts']]:
        eco_skipped.append((f'{plant}/{part}', 'the plant has no such part')); continue
    for mordant, colour in (('alum_potassium', alum), ('iron', iron)):
        if not colour:
            continue
        if make(plant, part, None, mordant, None, 'ecoprint', colour, '',
                'literature', ECO_SOURCE, quality=quality, note=note):
            eco_made += 1

if '--apply' in sys.argv:
    combos_pack['packVersion'] = '0.6.0'
    sources_pack['packVersion'] = 6
    json.dump(combos_pack, open('seed/combinations.json', 'w'), ensure_ascii=False, indent=1)
    json.dump(sources_pack, open('seed/sources.json', 'w'), ensure_ascii=False, indent=1)

print(f'FILLED    {len(filled)}   (empty → value on an existing record)')
print(f'CREATED   {len(created)}   new combination records ({eco_made} of them eco print)')
print(f'HELD      {len(held)}')
print()
for k, n in collections.Counter(f for _, f, _ in filled).most_common():
    print(f'  {k:16} {n}')
print()
for p, why in eco_skipped:
    print(f'  ECO SKIPPED  {p:32} {why}')
for h in held:
    print(f'  HELD  {str(h[0])[:44]:44} {h[1]:12} {h[2]!r} → {h[3]!r}')
if '--apply' not in sys.argv:
    print('\n(dry run — pass --apply to write)')
