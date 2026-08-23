#!/usr/bin/env python3
"""Rebuild the combination grid over the WHOLE key (§13ck).

The first grid asked for four things — part, mordant, pH, process — and the key
has eight. Filled in good faith, its rows could not have merged: three of the
missing four are part of what identifies a combination, so a row without them
does not name one record rather than naming it vaguely.

WHAT WAS MISSING, AND WHY EACH MATTERS

  fibreClass    Every one of the 28 existing records is cellulose. Cotton and
                wool take a different colour from the same plant, so a row that
                does not say which is not an answer about either.

  mordantBand   Combinations are matched by BAND, not by exact value — a fixed
                decision. „alum_potassium" without trace | low | medium | high
                is half the answer.

  blanket       This is the one that was actively losing information. Three rows
                describe a leaf printed with an IRON BLANKET, and they were left
                with an empty mordant on the correct reasoning that an iron
                blanket is not an iron mordant. But the key has a separate
                `blanket` field, and blank threw away a fact the record knew.
                They arrive pre-filled here.

  medium        AND THIS IS WHERE pH LIVES. The first pass reported that pH was
                not in the key at all; it is, under the name `medium`, as
                `{ phCode, whereCode }` — two of the 28 already carry
                `{alkaline, dye_bath}`. It was looked for by the wrong name.

                `whereCode` is what settles the outstanding question about
                safflower: pH OF THE BATH is `medium`, while an alkaline
                EXTRACTION is `extractionMode` on the dose (§13cc), and the two
                are different facts about different moments.

CARRIED FORWARD, NOT ASKED AGAIN. Everything already filled in the returned
workbook — 88 parts, 82 mordants, the processes, the notes — is written back into
this one. A grid that made someone answer the same 88 questions twice would not
be filled a second time.

Usage:  python3 scripts/export-combination-grid-2.py <returned.xlsx> [out.xlsx]
"""

import json
import pathlib
import re
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
DEFAULT_OUT = 'kombinacii-reshetka-2.xlsx'

INK, INDIGO, MADDER, LINE = '2A2724', '2C3B57', 'A03D3B', 'DED8CA'
SURFACE, GROUND, WELD_PALE = 'FFFDF8', 'F7F4EC', 'F7F1DC'

HEAD = Font(name='Arial', bold=True, size=10, color='FFFFFF')
BODY = Font(name='Arial', size=10, color=INK)
MUTED = Font(name='Arial', size=10, color='5C574E')
ITAL = Font(name='Arial', size=9, italic=True, color='5C574E')
NOTE_F = Font(name='Arial', size=10, color=MADDER, bold=True)
SUGGEST = Font(name='Arial', size=10, color=MADDER)

HEAD_FILL = PatternFill('solid', fgColor=INDIGO)
EDIT_FILL = PatternFill('solid', fgColor=SURFACE)
CTX_FILL = PatternFill('solid', fgColor=GROUND)
FLAG_FILL = PatternFill('solid', fgColor=WELD_PALE)

THIN = Side(style='thin', color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical='top', wrap_text=True)

# Read from vocab.js rather than restated: a second copy of a vocabulary drifts.
def vocab(dimension):
    src = (ROOT / 'vocab.js').read_text(encoding='utf-8')
    return [m.group(1) for m in
            re.finditer(rf"V\('{dimension}',\s*'([a-z_]+)'", src)]


def bands(dimension):
    src = (ROOT / 'vocab.js').read_text(encoding='utf-8')
    block = src.split('export const BANDS')[1]
    return [m.group(1) for m in
            re.finditer(rf"B\('{dimension}',\s*'([a-z_]+)'", block)] or \
           [m.group(2) for m in
            re.finditer(rf"'({dimension})',\s*'([a-z_]+)'", block)]


COLUMNS = [
    # (heading, width, editable)
    ('Растение',            22, False),
    ('Код',                 22, False),
    ('Цвят',                22, False),
    ('HEX',                  9, False),
    ('Условия (както са записани)', 40, False),
    ('Възможни части',      20, False),
    ('Част',                14, True),
    ('Влакно',              14, True),
    ('Байц',                18, True),
    ('Сила на байца',       15, True),
    ('Процес',              13, True),
    ('Одеяло',              14, True),
    ('pH на банята',        14, True),
    ('Къде е pH-то',        15, True),
    ('Сигурност',           13, False),
    ('Състояние',           13, True),
    ('Бележка',             40, True),
]
EDITABLE = {i for i, (_, _, e) in enumerate(COLUMNS, start=1) if e}

LEGEND = [
    ('РЕШЕТКА ЗА КОМБИНАЦИИТЕ — ВТОРА ВЕРСИЯ, ЦЕЛИЯТ КЛЮЧ.', NOTE_F),
    ('Първата питаше за четири неща, а ключът има осем. Тези редове не биха се слели.', BODY),
    ('Новото са четири колони: Влакно · Сила на байца · Одеяло · pH на банята (и къде е то).', BODY),
    ('Попълненото досега е пренесено — не се попълва втори път.', BODY),
    ('ВЛАКНО е задължително. Памук и вълна дават различен цвят от едно и също растение;', BODY),
    ('   ред без влакно не е отговор за нито едното. Всичките 28 съществуващи са целулоза.', BODY),
    ('СИЛА НА БАЙЦА: ключът се съпоставя по ЛЕНТИ, не по точни числа. „alum_potassium" сам е половин отговор.', BODY),
    ('ОДЕЯЛО: желязното одеяло НЕ е железен байц — прав е, който ги е разделил. Но одеялото е', BODY),
    ('   собствено поле, а празното изхвърляше верен факт. Трите такива реда идват попълнени.', BODY),
    ('pH: беше търсено под грешно име. В ключа е, като `medium`. „Къде" различава банята от екстракцията:', BODY),
    ('   алкална ЕКСТРАКЦИЯ не е pH на банята — тя е начин на извличане и стои на дозата.', BODY),
    ('Ако не се знае — празно. Редът просто няма да се слее; по-добре от грешен ключ.', ITAL),
]


def main():
    if len(sys.argv) < 2:
        sys.exit('usage: export-combination-grid-2.py <returned.xlsx> [out.xlsx]')
    src_path = pathlib.Path(sys.argv[1])
    out_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT

    src = load_workbook(src_path)['Комбинации']
    head = {str(src.cell(row=1, column=c).value or '').strip(): c
            for c in range(1, src.max_column + 1)}

    need = ['Растение', 'Код', 'Част', 'Възможни части', 'Цвят', 'HEX',
            'Условия (както са записани)', 'Байц', 'pH', 'Процес', 'Сигурност']
    missing = [h for h in need if h not in head]
    if missing:
        sys.exit('the returned workbook has no column: ' + ', '.join(missing))

    fibres = vocab('fibre_class')
    processes = vocab('process')
    wheres = vocab('medium_where')
    ph_codes = ['acid', 'neutral', 'alkaline']
    band_codes = ['trace', 'low', 'medium', 'high']

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('Комбинации')

    for i, (line, font) in enumerate(LEGEND, start=1):
        ws.cell(row=i, column=1, value=line).font = font
        ws.merge_cells(start_row=i, start_column=1,
                       end_row=i, end_column=len(COLUMNS))
        ws.row_dimensions[i].height = 14
    top = len(LEGEND) + 2

    for i, (h, w, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=top, column=i, value=h)
        cell.font, cell.fill, cell.border = HEAD, HEAD_FILL, BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[top].height = 30
    ws.freeze_panes = f'G{top + 1}'

    def get(r, name):
        v = src.cell(row=r, column=head[name]).value
        return '' if v is None else str(v).strip()

    # ONE OBSERVATION, TWO PARTS, TWO ROWS.
    #
    # „Дрян, жълто-кафяво, от листа и кора" is one line of the guide and two
    # records: the key takes a single part, so a row naming both names neither.
    # Leaving it blank would have thrown the observation away.
    #
    # Split rather than blanked, and each half carries a note saying it came
    # from a shared line — the colour was recorded for the two together, and
    # whether each part gives it alone is a question the split does not answer.
    SPLIT = {
        ('cornus_mas', 'от листа и кора'): ['лист', 'кора'],
    }

    row = top + 1
    blankets, ph_moved, carried, split_rows = 0, 0, 0, 0

    for r in range(2, src.max_row + 1):
        if not get(r, 'Растение'):
            continue

        conditions = get(r, 'Условия (както са записани)')
        note = get(r, 'Бележка') if 'Бележка' in head else ''

        # An iron blanket is not an iron mordant — and it is not nothing either.
        # Suggested only where the text says so plainly, and marked for checking.
        blanket = ''
        if re.search(r'желяз\w*\s+одеа?ло', conditions, re.I):
            blanket = 'iron'
            blankets += 1

        # pH was answered in the first grid against a column that had no home.
        # It has one — but only as the pH OF THE BATH, so it is carried over
        # with `dye_bath` and marked, never assumed for an extraction.
        ph = get(r, 'pH')
        where = ''
        if ph:
            where = 'dye_bath'
            ph_moved += 1

        parts_here = SPLIT.get((get(r, 'Код'), conditions.strip()))
        if parts_here:
            split_rows += len(parts_here) - 1

        for part in (parts_here or [get(r, 'Част')]):
            own_note = note
            if parts_here:
                own_note = (note + ' ' if note else '') + \
                    'От общ ред „' + conditions + '" — цветът е записан за двете части ' \
                    'заедно; дали всяка го дава сама е отделен въпрос.'

            vals = [
                get(r, 'Растение'), get(r, 'Код'), get(r, 'Цвят'), get(r, 'HEX'),
                conditions, get(r, 'Възможни части'),
                part, '',                             # fibre: nobody has said
                get(r, 'Байц'), '',                   # band: new
                get(r, 'Процес'), blanket, ph, where,
                get(r, 'Сигурност'),
                get(r, 'Състояние') if 'Състояние' in head else '',
                own_note,
            ]
            if any(vals[6:11]):
                carried += 1

            for i, v in enumerate(vals, start=1):
                # `None`, not `''`. openpyxl writes an empty string as a cell
                # that LOOKS empty and is not — it counts as filled, sorts oddly,
                # and a merge reading it back gets a value nobody typed.
                cell = ws.cell(row=row, column=i, value=(v or None))
                cell.border, cell.alignment = BORDER, WRAP
                if i in EDITABLE:
                    cell.font = BODY
                    cell.fill = EDIT_FILL
                    if not v:
                        cell.fill = FLAG_FILL
                else:
                    cell.font = MUTED
                    cell.fill = CTX_FILL
            # Suggested cells are marked, so a suggestion is never mistaken for
            # an answer someone gave.
            if blanket:
                ws.cell(row=row, column=12).font = SUGGEST
            if ph:
                ws.cell(row=row, column=14).font = SUGGEST
            if parts_here:
                ws.cell(row=row, column=7).font = SUGGEST
            ws.row_dimensions[row].height = 28
            row += 1

    guide = wb.create_sheet('Речник на стойностите')
    guide.column_dimensions['A'].width = 22
    guide.column_dimensions['B'].width = 78
    rows = [
        ('Колона', 'Позволени стойности'),
        ('Част', 'Само една. Ключът приема една част. „Лист и кора" са ДВА реда, не един.'),
        ('Влакно', ' · '.join(fibres) + '   — задължително'),
        ('Байц', 'none · alum_potassium · alum_acetate · iron · copper · tannin'),
        ('Сила на байца', ' · '.join(band_codes) + '   — лента, не число'),
        ('Процес', ' · '.join(processes)),
        ('Одеяло', 'iron · dye · друг материал · празно, ако няма одеяло'),
        ('pH на банята', ' · '.join(ph_codes) + '   — празното НЕ значи neutral'),
        ('Къде е pH-то', ' · '.join(wheres)),
        ('Състояние', 'draft · review · ok'),
        ('', ''),
        ('Не се попълва', 'Алкална ЕКСТРАКЦИЯ не е pH. Тя е начин на извличане и стои на дозата.'),
        ('', 'Наслагване върху друго багрило не е комбинация — ключът не описва два цвята.'),
        ('', 'Сила на банята не е измерение. „Слаба баня" е бележка, не байц.'),
    ]
    for i, (a, b) in enumerate(rows, start=1):
        guide.cell(row=i, column=1, value=a).font = NOTE_F if i == 1 else BODY
        guide.cell(row=i, column=2, value=b).font = HEAD if i == 1 else MUTED
        for c in (1, 2):
            guide.cell(row=i, column=c).alignment = WRAP
        guide.row_dimensions[i].height = 20

    wb.save(out_path)
    n = row - top - 1
    print(f'{out_path}: {n} rows over the whole key')
    print(f'  carried forward from the returned grid : {carried} rows')
    print(f'  iron blankets recovered                : {blankets}')
    print(f'  pH answers rehoused as bath pH         : {ph_moved}')
    print(f'  rows added by splitting a shared line   : {split_rows}')
    print(f'  new and empty for everyone             : Влакно, Сила на байца')


if __name__ == '__main__':
    main()
