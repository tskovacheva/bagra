#!/usr/bin/env python3
"""Merge the filled plant workbook back into `seed/plants.json` (§13ce).

Same discipline as the chemistry merge: it does not overwrite what is already
there, it does not invent, and it stops at anything it cannot read rather than
choosing an interpretation.

Re-runnable. A second run reports nothing to do.

WHAT COMES BACK

1.  A GENERAL DESCRIPTION for all 57 plants, in both languages — `description`.

    The plant as a plant, before it is a dye. This is NOT `character` and not the
    „Как се държи“ section; both of those describe dye behaviour — the temperament
    of the bath, what spoils it, the order of work (§13m). A botanical opening is
    a seventh thing, and by the rule the library rests on — what must be present
    on every plant is a field, not a section — it is a field.

2.  GATHERING MONTHS PER PART — `part.harvestMonths`, 111 rows.

    The third instance of one fault, after the temperature on the plant (§13az)
    and the extraction mode on the part (§13cc): a value fixed to a record that
    does not determine it. Walnut leaf is May–September and the green husks are
    August–October; one list per plant could hold one of those and made the other
    unsayable.

3.  SEVEN PARTS THAT ARE BOUGHT, NOT GATHERED — `part.sourcedNotGathered`.

    Avocado stone and skin, cutch bark and heartwood, henna leaf, brazilwood bark
    and heartwood. A positive statement, distinct from „nobody has filled this in
    yet“: a month of harvest for cutch would have been invented, and an invented
    reason is the same fault with better grammar.

    Note what did NOT turn out to be in this group. Safflower, sumac, pomegranate,
    Persian buckthorn and eucalyptus are marked `habitat: imported` and came back
    WITH months. That vocabulary is wild | garden | imported and answers where a
    plant GROWS; sumac is native here and also arrives in a bag, and both are true
    at once. Reading the origin field as a sourcing field was a misreading, not a
    contradiction in the data — two facts had been sharing one field.

WHAT IT DOES NOT TOUCH

`plant.harvestMonths` stays. It is superseded, not deleted, and it is retired one
version later — as `FabricStateEvent` was (§13bd) — so the new per-part answers
can be checked against real records first. `modules/season.js` already prefers the
part and carries the fallback in the open.

`character` and the „Как се държи“ section both stay as they are. They carry the
same heading and appear twice on fourteen records, which wants deciding, but it is
a different decision and merging a third prose block is not the moment to take it.

Usage:  python3 scripts/merge-plant-gaps.py <workbook.xlsx> [--apply]

Without `--apply` it reports and writes nothing.
"""

import json
import pathlib
import re
import sys

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parent.parent
PLANTS = ROOT / 'seed' / 'plants.json'
VOCAB = ROOT / 'vocab.js'

PACK_VERSION = '0.8.0'
NOT_GATHERED = 'не се бере'
FILL_ME = 'ЗА ПОПЪЛВАНЕ'

MONTHS_BG = ['януари', 'февруари', 'март', 'април', 'май', 'юни',
             'юли', 'август', 'септември', 'октомври', 'ноември', 'декември']

# The months in this pack were observed in Bulgaria. A gathering month is not a
# property of the plant; it is a property of the plant HERE — walnut husks in
# Sofia and in Andalusia are not the same week, and the library is a candidate
# for distribution. Not climate zones: one label, so that a second region can be
# added later instead of 118 rows having to be guessed at (§13cd).
HARVEST_REGION = 'BG'

HEADER_ROW = 7          # the legend stands above the table in both sheets
FIRST_DATA_ROW = 8


def part_labels():
    """`bg label -> partCode`, read from vocab.js rather than restated here."""
    src = VOCAB.read_text(encoding='utf-8')
    out = {}
    for m in re.finditer(r"V\('plant_part',\s*'([a-z_]+)',\s*'([^']*)'", src):
        out[m.group(2).strip().lower()] = m.group(1)
    return out


def parse_months(cell, where, problems):
    """„юни · юли · август“ -> [6, 7, 8].

    Returns None for „не се бере“ and for anything unreadable — with the reason
    recorded. A month list that half-parses is worse than one that fails: it
    would put the plant on the seasonal panel for the months it happened to
    understand.
    """
    raw = (cell or '').strip()
    if not raw or raw == FILL_ME:
        return ('unfilled', None)
    if raw.lower() == NOT_GATHERED:
        return ('bought', None)

    months = []
    for token in re.split(r'[·,;/]+', raw):
        token = token.strip().lower()
        if not token:
            continue
        if token not in MONTHS_BG:
            problems.append(f'{where}: cannot read „{token}“ as a month')
            return ('bad', None)
        months.append(MONTHS_BG.index(token) + 1)

    if not months:
        problems.append(f'{where}: nothing readable in „{raw}“')
        return ('bad', None)
    return ('months', sorted(set(months)))


def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__.strip().splitlines()[-3])
    book = pathlib.Path(sys.argv[1])
    apply = '--apply' in sys.argv

    wb = load_workbook(book)
    data = json.loads(PLANTS.read_text(encoding='utf-8'))
    plants = data['plants']

    by_name = {}
    for p in plants:
        by_name.setdefault((p['nameCommon'].get('bg') or '').strip(), []).append(p)

    partcode = part_labels()
    problems = []

    # ---------------------------------------------------------------- 1. prose
    ws = wb['Описание']
    desc_new, desc_kept = [], []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        name = (ws.cell(row=r, column=1).value or '').strip()
        if not name:
            continue
        found = by_name.get(name)
        if not found:
            problems.append(f'Описание r{r}: no plant called „{name}“')
            continue
        if len(found) > 1:
            problems.append(f'Описание r{r}: „{name}“ matches {len(found)} plants')
            continue
        p = found[0]

        bg = (ws.cell(row=r, column=6).value or '').strip()
        en = (ws.cell(row=r, column=7).value or '').strip()
        if not bg or bg == FILL_ME:
            continue
        if not en:
            problems.append(f'Описание r{r} ({name}): Bulgarian filled, English empty — '
                            'the library is bilingual from the first record')
            continue

        existing = (p.get('description') or {}).get('bg', '').strip()
        if existing and existing != bg:
            # Never overwritten. If a description is already there and differs,
            # something has been edited on one side or the other, and choosing
            # between them is not this script's decision.
            desc_kept.append(name)
            continue
        if existing == bg:
            continue
        p['description'] = {'bg': bg, 'en': en}
        desc_new.append(name)

    # ------------------------------------------------------------- 2. gathering
    ws = wb['Беритба']
    months_set, bought_set, months_kept, unfilled = [], [], [], []
    seen_pairs = set()

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        name = (ws.cell(row=r, column=1).value or '').strip()
        partname = (ws.cell(row=r, column=3).value or '').strip()
        if not name:
            continue

        found = by_name.get(name)
        if not found or len(found) > 1:
            problems.append(f'Беритба r{r}: „{name}“ matches '
                            f'{len(found) if found else 0} plants')
            continue
        p = found[0]

        code = partcode.get(partname.lower())
        if code is None and partname not in ('—', ''):
            problems.append(f'Беритба r{r}: „{partname}“ is not a plant part')
            continue

        part = next((x for x in p.get('parts', []) if x.get('partCode') == code), None)
        if part is None:
            problems.append(f'Беритба r{r}: {p["code"]} has no part „{partname}“')
            continue

        key = (p['code'], code)
        if key in seen_pairs:
            problems.append(f'Беритба r{r}: {p["code"]}.{code} appears twice')
            continue
        seen_pairs.add(key)

        where = f'Беритба r{r} ({name}, {partname})'
        kind, months = parse_months(ws.cell(row=r, column=6).value, where, problems)
        note = (ws.cell(row=r, column=7).value or '').strip()

        if kind == 'bad':
            continue
        if kind == 'unfilled':
            unfilled.append(f'{p["code"]}.{code}')
            continue

        if kind == 'bought':
            # A part cannot be both bought and gathered. If months are already
            # recorded on it, that is a disagreement to be looked at, not
            # resolved by whichever ran last.
            if part.get('harvestMonths'):
                problems.append(f'{where}: marked as bought but already carries months')
                continue
            # Reported only when it CHANGES. A count that includes what was
            # already there makes a second run look like a second write, and a
            # report nobody can trust is a report nobody reads.
            if not part.get('sourcedNotGathered'):
                bought_set.append(f'{p["code"]}.{code}')
            part['sourcedNotGathered'] = True
            part.pop('harvestMonths', None)
            continue

        existing = part.get('harvestMonths')
        if existing and existing != months:
            months_kept.append(f'{p["code"]}.{code}')
            continue
        if existing == months:
            continue
        if part.get('sourcedNotGathered'):
            problems.append(f'{where}: has months but is marked as bought')
            continue
        part['harvestMonths'] = months
        if note:
            part['harvestNote'] = {'bg': note, 'en': ''}
        months_set.append(f'{p["code"]}.{code}')

    # -------------------------------------------------------------- 3. checks
    for p in plants:
        for part in p.get('parts', []):
            if part.get('harvestMonths') == []:
                problems.append(f'{p["code"]}.{part.get("partCode")}: empty list — '
                                '„gathered in no month" is not true of any part; use no field')
            for m in part.get('harvestMonths') or []:
                if not isinstance(m, int) or not 1 <= m <= 12:
                    problems.append(f'{p["code"]}.{part.get("partCode")}: „{m}“ is not a month')

    # ------------------------------------------------------------- 4. report
    covered = sum(1 for p in plants for x in p.get('parts', [])
                  if x.get('harvestMonths') or x.get('sourcedNotGathered'))
    total = sum(len(p.get('parts', [])) for p in plants)
    described = sum(1 for p in plants if (p.get('description') or {}).get('bg'))

    print(f'{book.name}')
    print(f'  descriptions written   : {len(desc_new)}  (library now {described}/{len(plants)})')
    print(f'  months written         : {len(months_set)}')
    print(f'  marked bought          : {len(bought_set)}  ({", ".join(bought_set)})')
    print(f'  parts now answered     : {covered}/{total}')
    if desc_kept:
        print(f'  description already there and different, kept: {", ".join(desc_kept)}')
    if months_kept:
        print(f'  months already there and different, kept: {", ".join(months_kept)}')
    if unfilled:
        print(f'  still unfilled ({len(unfilled)}): {", ".join(unfilled)}')

    if problems:
        print(f'\n  {len(problems)} thing(s) this script will not decide:')
        for x in problems:
            print(f'    - {x}')
        sys.exit('stopped: nothing written')

    if not apply:
        print('\n  dry run — pass --apply to write')
        return

    data['packVersion'] = PACK_VERSION
    data['harvestRegion'] = HARVEST_REGION
    PLANTS.write_text(json.dumps(data, ensure_ascii=False, indent=1) + '\n',
                      encoding='utf-8')
    print(f'\n  written · pack {PACK_VERSION} · harvestRegion {HARVEST_REGION}')


if __name__ == '__main__':
    main()
