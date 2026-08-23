#!/usr/bin/env python3
"""Export the two gaps in the plant library for the owner to fill (§13cc).

WHAT IS MISSING, AND WHY IT IS TWO SHEETS RATHER THAN ONE

1.  NO PLANT HAS A GENERAL DESCRIPTION.  The record opens straight into dye
    figures. What is wanted is the opening paragraph a reference book gives —
    „Брошът е едно от най-древните багрилни растения, многогодишно тревисто,
    виреещо…" — the plant as a plant, before the plant as a dye.

    This is NOT `character` and it is not the „Как се държи" section. Both of
    those describe dye behaviour: the temperament of the bath, what spoils it,
    the order of work (§13m). A botanical opening is a seventh thing, and by the
    rule the library already rests on — what must be present on every plant is a
    FIELD, not a section — it is a field.

    Sheet `Описание`, one row per plant, 57 rows, all 57 empty.

2.  HARVEST MONTHS ARE MISSING ON THIRTEEN PLANTS — AND THE THIRTEEN ARE NOT
    ONE PROBLEM.

    Nine are simply unfilled: the nine plants added most recently, which also
    lack `character`. Those want filling.

    Four are `imported` — brazilwood, cutch, henna, avocado. These are bought,
    not gathered. A month of harvest for cutch would be invented, and an
    invented reason is the same fault with better grammar. Empty and
    NOT-APPLICABLE are different statements and the data cannot currently tell
    them apart, exactly as the chemistry audit could not tell an unrecorded
    strength from an absent compound until it was given a word for it. The sheet
    therefore offers `не се бере` as a value, not a blank.

    AND THE FIELD IS ON THE WRONG RECORD. `harvestMonths` sits on the PLANT, but
    parts are not gathered together: ash is leaf and bark, elder is leaf, bark
    and fruit — three separate windows. Bark in winter when the sap is down,
    leaf in summer, fruit in autumn. One set of months per plant cannot say that,
    and where it is filled today it is presumably the months of whichever part
    was in mind at the time, with no way to know which.

    This is the same shape as the `extractionMode` fault (§13bv): a value fixed
    to a record that does not determine it.

    So the sheet is ONE ROW PER PART, not per plant. That is the safe direction
    regardless of what is decided about the model: part-level answers can always
    be collapsed to a plant-level field, and plant-level answers can never be
    expanded into parts. Filling this in cannot be wasted work.

WHAT THIS SCRIPT DOES NOT DO

It does not change the model. No field is added, nothing is migrated. It reads
`seed/plants.json` and writes a workbook. The merge back is a separate script
with the same discipline as the chemistry merge — refuses to overwrite, refuses
to guess, stops at anything unclear.

Re-runnable. Reads only.

Usage:  python3 scripts/export-plant-gaps.py [output.xlsx]
"""
import json
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PLANTS = 'seed/plants.json'
VOCAB = 'vocab.js'
DEFAULT_OUT = 'plant-gaps.xlsx'

FILL_ME = 'ЗА ПОПЪЛВАНЕ'
NOT_HARVESTED = 'не се бере'

MONTHS_BG = ['януари', 'февруари', 'март', 'април', 'май', 'юни',
             'юли', 'август', 'септември', 'октомври', 'ноември', 'декември']

# Palette (§ palette). No green anywhere, including here: the workbook is read
# beside the swatches and a green header would sit in the eye while a colour is
# being judged.
INK = '2A2724'
INDIGO = '2C3B57'
MADDER = 'A03D3B'
LINE = 'DED8CA'
SURFACE = 'FFFDF8'
GROUND = 'F7F4EC'
WELD_PALE = 'F7F1DC'

HEAD = Font(name='Arial', bold=True, size=10, color='FFFFFF')
BODY = Font(name='Arial', size=10, color=INK)
BODY_MUTED = Font(name='Arial', size=10, color='5C574E')
BODY_IT = Font(name='Arial', size=9, italic=True, color='5C574E')
NOTE = Font(name='Arial', size=10, color=MADDER, bold=True)

HEAD_FILL = PatternFill('solid', fgColor=INDIGO)
EDIT_FILL = PatternFill('solid', fgColor=SURFACE)
CTX_FILL = PatternFill('solid', fgColor=GROUND)
FLAG_FILL = PatternFill('solid', fgColor=WELD_PALE)

THIN = Side(style='thin', color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical='top', wrap_text=True)
TOP = Alignment(vertical='top')


def vocab_pairs(kind):
    """`code -> bg` for one vocabulary, read from vocab.js.

    Read rather than restated: a second copy of the vocabulary is a second thing
    to keep in step, and it would drift.
    """
    src = open(VOCAB, encoding='utf-8').read()
    out = {}
    for m in re.finditer(rf"'{kind}',\s*'([a-z_]+)',\s*'([^']*)',\s*'([^']*)'", src):
        code, bg = m.group(1), m.group(2)
        if re.search(r'[а-яА-Я]', bg):
            out[code] = bg
    return out


def text_of(field, lang='bg'):
    if isinstance(field, dict):
        return field.get(lang) or ''
    return field or ''


def months_text(nums):
    return ' · '.join(MONTHS_BG[n - 1] for n in sorted(nums or []))


def style_row(ws, row, ncols, editable_cols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = BORDER
        cell.alignment = TOP_WRAP
        if c in editable_cols:
            cell.font = BODY
            cell.fill = EDIT_FILL
        else:
            cell.font = BODY_MUTED
            cell.fill = CTX_FILL


def write_header(ws, headers, widths, freeze):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = freeze


def sheet_legend(ws, lines, ncols):
    """A note above the table saying which cells are meant to be typed in.

    Above rather than on another tab: a legend on its own sheet is a legend
    nobody reads.
    """
    for i, (line, font) in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line).font = font
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=ncols)
        ws.row_dimensions[i].height = 15
    ws.row_dimensions[len(lines) + 1].height = 6
    return len(lines) + 2


def build_description(wb, plants, types, habitats):
    ws = wb.create_sheet('Описание')
    headers = ['Растение', 'Латинско име', 'Семейство', 'Вид', 'Местообитание',
               'ОБЩО ОПИСАНИЕ (BG) — попълни', 'GENERAL DESCRIPTION (EN) — попълни',
               'Как се държи (има го вече — за справка)']
    widths = [26, 24, 16, 14, 18, 62, 62, 50]
    ncols = len(headers)

    top = sheet_legend(ws, [
        ('ОБЩО ОПИСАНИЕ НА РАСТЕНИЕТО — 57 реда, всичките празни.', NOTE),
        ('Растението като растение, преди да е багрило: какво е, откъде е, къде вирее, '
         'за какво е известно. Две-три изречения.', BODY),
        ('Попълват се само двете бели колони. Останалото е контекст.', BODY),
        ('Последната колона е СЪЩЕСТВУВАЩИЯТ текст за багрилното поведение — той е друго '
         'нещо и не се преписва тук.', BODY_IT),
        ('Пример: „Брошът е многогодишно тревисто растение от семейство Брошови, родом от '
         'Предна Азия и Средиземноморието. Отглежда се от древността заради корените си, '
         'които са най-старият познат източник на трайно червено багрило."', BODY_IT),
    ], ncols)

    # The header cannot be row 1 once a legend stands above it; openpyxl's
    # freeze_panes takes the first scrolling cell, so it is the row below.
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=top, column=i, value=h)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[top].height = 30
    ws.freeze_panes = f'A{top + 1}'

    row = top + 1
    for p in sorted(plants, key=lambda x: text_of(x['nameCommon'])):
        ws.cell(row=row, column=1, value=text_of(p['nameCommon']))
        ws.cell(row=row, column=2, value=p.get('nameBotanical', ''))
        ws.cell(row=row, column=3, value=p.get('family', ''))
        ws.cell(row=row, column=4, value=types.get(p.get('plantType'), p.get('plantType') or ''))
        ws.cell(row=row, column=5,
                value=' · '.join(habitats.get(h, h) for h in (p.get('habitat') or [])))
        ws.cell(row=row, column=6, value=FILL_ME)
        ws.cell(row=row, column=7, value='')
        # Whichever of the two already says something about behaviour, so she can
        # see what is already said and not repeat it.
        existing = text_of(p.get('character'))
        if not existing:
            for s in p.get('sections', []):
                if s['title']['bg'] == 'Как се държи':
                    existing = text_of(s['body'])
        ws.cell(row=row, column=8, value=existing)
        style_row(ws, row, ncols, editable_cols={6, 7})
        ws.cell(row=row, column=6).font = Font(name='Arial', size=10, color=MADDER)
        ws.cell(row=row, column=6).fill = FLAG_FILL
        ws.row_dimensions[row].height = 58
        row += 1
    return row - top - 1


def build_harvest(wb, plants, parts_bg, habitats):
    ws = wb.create_sheet('Беритба')
    headers = ['Растение', 'Латинско име', 'Част', 'Местообитание',
               'Месеци сега (на РАСТЕНИЕТО)', 'МЕСЕЦИ ЗА ТАЗИ ЧАСТ — попълни',
               'Бележка за беритбата (по избор)']
    widths = [26, 22, 16, 18, 30, 40, 46]
    ncols = len(headers)

    top = sheet_legend(ws, [
        ('МЕСЕЦИ НА БЕРИТБА — един ред за ВСЯКА ЧАСТ, не за растение.', NOTE),
        ('Кората и листът на един и същ храст не се берат по едно и също време, а полето '
         'в момента е едно за цялото растение. Затова питаме по части.', BODY),
        ('Пиши имената на месеците, разделени с „·" — например: юни · юли · август', BODY),
        (f'Ако частта не се бере, защото се купува (внос, екстракт), напиши точно: '
         f'{NOT_HARVESTED}. Не оставяй празно — празното значи „още не сме стигнали дотам".', BODY),
        ('Петата колона показва какво стои днес на растението. То може да е вярно само за '
         'една от частите.', BODY_IT),
    ], ncols)

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=top, column=i, value=h)
        cell.font = HEAD
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[top].height = 30
    ws.freeze_panes = f'A{top + 1}'

    row = top + 1
    n_rows = 0
    for p in sorted(plants, key=lambda x: text_of(x['nameCommon'])):
        current = months_text(p.get('harvestMonths'))
        habitat = p.get('habitat') or []
        # `imported` alone is not the answer. Nine plants are marked imported and
        # five of them ALREADY carry months — sumac, pomegranate, safflower,
        # eucalyptus, buckthorn all grow here as well as arriving in a bag. To
        # pre-fill „не се бере" over a real recorded month would be substituting
        # a guess for data, which is the fault this workbook exists to end.
        # Suggested only where the plant is imported AND nothing is recorded.
        bought = 'imported' in habitat and not p.get('harvestMonths')
        # Imported WITH months is a contradiction in the record, not a gap. It is
        # shown as a question rather than answered here.
        contradiction = 'imported' in habitat and bool(p.get('harvestMonths'))
        # A plant with no parts still gets one row, so that it is not silently
        # absent from a sheet whose purpose is completeness.
        entries = p.get('parts') or [{'partCode': ''}]
        for part in entries:
            ws.cell(row=row, column=1, value=text_of(p['nameCommon']))
            ws.cell(row=row, column=2, value=p.get('nameBotanical', ''))
            ws.cell(row=row, column=3,
                    value=parts_bg.get(part.get('partCode'), part.get('partCode') or '—'))
            ws.cell(row=row, column=4,
                    value=' · '.join(habitats.get(h, h) for h in habitat))
            ws.cell(row=row, column=5, value=current or '(празно)')
            ws.cell(row=row, column=6, value=NOT_HARVESTED if bought else FILL_ME)
            ws.cell(row=row, column=7,
                    value='Означено е внос, но има записани месеци — кое от двете е вярно?'
                          if contradiction else '')
            style_row(ws, row, ncols, editable_cols={6, 7})
            c6 = ws.cell(row=row, column=6)
            if not bought:
                c6.font = Font(name='Arial', size=10, color=MADDER)
                c6.fill = FLAG_FILL
            if contradiction:
                ws.cell(row=row, column=7).font = NOTE
            ws.row_dimensions[row].height = 30
            row += 1
            n_rows += 1

    # A list, not a restriction: months are written freehand and a validation
    # that refuses „юни · юли" would fight the format the sheet asks for.
    dv = DataValidation(type='list', formula1=f'"{NOT_HARVESTED}"', allow_blank=True,
                        showErrorMessage=False)
    dv.prompt = ('Месеци, разделени с „·". Ако частта се купува, а не се бере: '
                 + NOT_HARVESTED)
    dv.promptTitle = 'Беритба'
    ws.add_data_validation(dv)
    dv.add(f'F{top + 1}:F{row - 1}')
    return n_rows


def build_readme(wb, n_desc, n_harv, imported_codes, unfilled_codes):
    ws = wb.create_sheet('Прочети', 0)
    ws.column_dimensions['A'].width = 110
    lines = [
        ('БАГРА — двете празнини в библиотеката с растения', NOTE),
        ('', BODY),
        (f'Лист „Описание" — {n_desc} реда, по един на растение. Всичките са празни: '
         'нито едно растение няма общо описание.', BODY),
        ('   Това е растението като растение, преди да е багрило. Не е „Как се държи" — '
         'то описва поведението на банята и вече съществува.', BODY_MUTED),
        ('', BODY),
        (f'Лист „Беритба" — {n_harv} реда, по един на ЧАСТ.', BODY),
        ('   Месеците днес стоят на растението, а кората и листът не се берат заедно. '
         'Затова се пита по части: така отговорът върши работа и ако полето остане на '
         'растението, и ако се премести на частта.', BODY_MUTED),
        (f'   {len(unfilled_codes)} растения нямат никакви месеци, защото са добавени '
         'последни и не са попълвани.', BODY_MUTED),
        (f'   {len(imported_codes)} са внос — купуват се, не се берат. Техните редове са '
         f'предварително попълнени с „{NOT_HARVESTED}". Празно и „не се прилага" са две '
         'различни твърдения и данните не бива да ги смесват.', BODY_MUTED),
        ('', BODY),
        ('Попълват се само белите колони. Оцветените в жълто са тези, които чакат теб.', BODY),
        ('Останалите колони са контекст, за да не се отваря приложението за всеки ред.', BODY_MUTED),
        ('', BODY),
        ('Когато върнеш файла: сливането минава през скрипт със същата дисциплина като '
         'химичния одит — не презаписва вече попълнено, не измисля, спира при неясно.', BODY),
    ]
    for i, (line, font) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = font
        c.alignment = Alignment(vertical='top', wrap_text=True)
        ws.row_dimensions[i].height = 28 if line else 8


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    plants = json.load(open(PLANTS, encoding='utf-8'))['plants']

    parts_bg = vocab_pairs('plant_part')
    types = vocab_pairs('plant_type')
    habitats = vocab_pairs('habitat')

    imported = [p['code'] for p in plants
                if 'imported' in (p.get('habitat') or []) and not p.get('harvestMonths')]
    contradictory = [p['code'] for p in plants
                     if 'imported' in (p.get('habitat') or []) and p.get('harvestMonths')]
    unfilled = [p['code'] for p in plants
                if not p.get('harvestMonths') and 'imported' not in (p.get('habitat') or [])]

    wb = Workbook()
    wb.remove(wb.active)
    n_desc = build_description(wb, plants, types, habitats)
    n_harv = build_harvest(wb, plants, parts_bg, habitats)
    build_readme(wb, n_desc, n_harv, imported, unfilled)
    wb.save(out_path)

    print(f'{out_path}')
    print(f'  Описание: {n_desc} rows, all empty')
    print(f'  Беритба:  {n_harv} rows across {len(plants)} plants')
    print(f'    {len(unfilled)} plants have no months and are not imported: '
          f'{", ".join(unfilled)}')
    print(f'    {len(imported)} imported with nothing recorded, suggested „{NOT_HARVESTED}": '
          f'{", ".join(imported)}')
    print(f'    {len(contradictory)} marked imported but DO carry months — asked, not answered: '
          f'{", ".join(contradictory)}')


if __name__ == '__main__':
    main()
