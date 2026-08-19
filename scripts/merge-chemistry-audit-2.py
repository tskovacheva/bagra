#!/usr/bin/env python3
"""Merge the second chemistry audit into seed/plants.json (1.0.0-rc8).

The first audit script — merge-chemistry-audit.py — deliberately refused three
things: creating an entry, removing one, and changing a compound's class. Those
are domain decisions and it had no standing to make them. This audit asks for
all three, so it gets a script that does them EXPLICITLY rather than smuggling
them through the old one.

What this script still refuses:

- **Overwriting a level that is already set.** Checked, not assumed: across the
  134 rows the audit and the library share, it overwrites nothing and blanks
  nothing, and all 19 entries it removes or reclasses have an empty level in
  the library. If that ever stops being true the run stops.
- **Removing an entry the audit does not explain.** `За merge` is a replacing
  set, so anything absent from it disappears. Read literally that makes a row
  dropped by accident indistinguishable from one dropped on purpose. Every
  disappearance must therefore be named in `Корекции вещества` or carry an
  empty level; an unexplained one halts the run.
- **Turning an empty level into a guess.** The audit distinguishes *not
  recorded* from *no honest quantitative estimate exists* via its `Статус на
  знанието` column. That lands on the entry as `levelUnknown: true`.

  NOT as `confidence: 'unknown'`, which is what the first draft of this script
  wrote. `confidence` is already a dimension in vocab.js with five values —
  unverified, literature, practice, confirmed, contradicted — and `unknown` is
  not one of them. Writing it there would have put an unknown code into a
  controlled vocabulary, which renders as its own key on screen and is exactly
  the fault guard 24c was written for. A separate boolean says the separate
  thing.

Dry by default. Pass --write to save.

Usage:
    python3 scripts/merge-chemistry-audit-2.py <workbook.xlsx>
    python3 scripts/merge-chemistry-audit-2.py <workbook.xlsx> --write
"""
import json
import sys
import re
import collections

from openpyxl import load_workbook

PLANTS = 'seed/plants.json'
VOCAB = 'vocab.js'
LEVELS = {'trace', 'moderate', 'high', 'dominant'}

# The audit renames one label. IUPAC separates anthocyanins (the glycosides)
# from anthocyanidins (their aglycones), and Bulgarian scientific usage follows
# with „антоцианини" / „антоцианидини". The library's own prose already said
# „антоцианини" in seed/plants.json, the glossary and the techniques, while
# vocab.js alone said „антоциани" — so this is internal consistency as much as
# correctness. The CODE (`anthocyanin`) does not change; only the label does.
#
# It matters here beyond terminology: until vocab.js is corrected, the six audit
# rows written as „антоцианини" resolve to no code at all and would be silently
# skipped.
RENAMED = {'антоцианини': 'anthocyanin'}


def vocab_pairs(kind):
    src = open(VOCAB, encoding='utf-8').read()
    out = {}
    for m in re.finditer(rf"'{kind}',\s*'([a-z_]+)',\s*'([^']*)',\s*'([^']*)'", src):
        code, bg = m.group(1), m.group(2)
        if re.search(r'[а-яА-Я]', bg):
            out[bg] = code
    return out


def read_sheet(wb, name, cols):
    ws = wb[name]
    rows = []
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(row=r, column=c).value for c in range(1, len(cols) + 1)]
        if not any(v):
            continue
        rows.append(dict(zip(cols, v)))
    return rows


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path, write = sys.argv[1], '--write' in sys.argv

    chem = vocab_pairs('chemistry_class')
    chem.update(RENAMED)
    part = vocab_pairs('plant_part')

    wb = load_workbook(path, data_only=True)

    # ---- the replacing set
    audit = {}
    for row in read_sheet(wb, 'За merge',
                          ['plant', 'bot', 'code', 'part', 'compound', 'level', 'status', 'note']):
        if not row['code']:
            continue
        pc = part.get(row['part'], row['part'])
        cc = chem.get(row['compound'])
        if cc is None:
            print(f"СПИРАМ: «{row['compound']}» не е клас вещество в речника "
                  f"({row['code']} / {row['part']})")
            sys.exit(1)
        audit[(row['code'], pc, cc)] = {
            'level': (row['level'] or '').strip(),
            'status': (row['status'] or '').strip(),
        }

    # ---- what the audit explains, used as permission to remove
    explained = set()
    for row in read_sheet(wb, 'Корекции вещества',
                          ['n', 'plant', 'part', 'before', 'after', 'level', 'action', 'why']):
        if row['plant'] and row['part'] and row['before']:
            explained.add((str(row['plant']).strip(), str(row['part']).strip(),
                           str(row['before']).strip()))

    # ---- the library
    doc = json.load(open(PLANTS, encoding='utf-8'), object_pairs_hook=collections.OrderedDict)
    plant_bg = {}
    base = {}
    for p in doc['plants']:
        nm = p.get('nameCommon') or {}
        plant_bg[p['code']] = (nm.get('bg') if isinstance(nm, dict) else nm) or p['code']
        for pt in (p.get('parts') or []):
            for c in (pt.get('chemistry') or []):
                base[(p['code'], pt.get('partCode'), c.get('classCode'))] = c

    part_bg = {v: k for k, v in part.items()}
    chem_bg = {v: k for k, v in vocab_pairs('chemistry_class').items()}

    # ---- the refusals, all checked before anything is written
    overwrite, blanked, unexplained = [], [], []

    for key in set(base) & set(audit):
        have = base[key].get('level') or ''
        want = audit[key]['level']
        if have and want and have != want:
            overwrite.append((key, have, want))
        if have and not want:
            blanked.append((key, have))

    for key in set(base) - set(audit):
        code, pc, cc = key
        have = base[key].get('level') or ''
        named = (plant_bg.get(code, code), part_bg.get(pc, pc), chem_bg.get(cc, cc))
        if named not in explained and have:
            unexplained.append((key, have))

    stop = False
    if overwrite:
        stop = True
        print(f'ОТКАЗ — презаписване на преценена степен ({len(overwrite)}):')
        for k, a, b in overwrite:
            print(f'   {k}: «{a}» -> «{b}»')
    if blanked:
        stop = True
        print(f'ОТКАЗ — изпразване на преценена степен ({len(blanked)}):')
        for k, a in blanked:
            print(f'   {k}: «{a}»')
    if unexplained:
        stop = True
        print(f'ОТКАЗ — премахване без обяснение ({len(unexplained)}):')
        for k, a in unexplained:
            print(f'   {k}: «{a}» — няма го нито в «За merge», нито в «Корекции вещества»')
    if stop:
        print('\nНищо не е записано.')
        sys.exit(1)

    # ---- apply
    added = filled = removed = marked = 0
    for p in doc['plants']:
        for pt in (p.get('parts') or []):
            pc = pt.get('partCode')
            chemistry = pt.get('chemistry') or []

            keep = []
            for c in chemistry:
                key = (p['code'], pc, c.get('classCode'))
                if key not in audit:
                    removed += 1
                    continue
                a = audit[key]
                if a['level'] and not (c.get('level') or ''):
                    c['level'] = a['level']
                    filled += 1
                # A statement, not a gap: no honest quantitative estimate
                # exists, usually because the plant is strongly seasonal or
                # cultivar-dependent. An empty level without this mark still
                # means simply "not recorded".
                if a['status'] == 'unknown':
                    c['levelUnknown'] = True
                    marked += 1
                keep.append(c)

            for (code, apc, acc), a in audit.items():
                if code != p['code'] or apc != pc:
                    continue
                if any(c.get('classCode') == acc for c in keep):
                    continue
                entry = collections.OrderedDict({'classCode': acc})
                if a['level']:
                    entry['level'] = a['level']
                else:
                    entry['level'] = ''
                if a['status'] == 'unknown':
                    entry['levelUnknown'] = True
                    marked += 1
                keep.append(entry)
                added += 1

            if keep or chemistry:
                pt['chemistry'] = keep

    print(f'добавени записи:      {added}')
    print(f'попълнени степени:    {filled}')
    print(f'премахнати записи:    {removed}')
    print(f'маркирани «неизвестна»: {marked}')

    total = sum(len(pt.get('chemistry') or [])
                for p in doc['plants'] for pt in (p.get('parts') or []))
    print(f'общо химични записи:  {total}')

    if not write:
        print('\n(пробен пуск — нищо не е записано. Пусни с --write.)')
        return

    with open(PLANTS, 'w', encoding='utf-8') as f:
        json.dump(doc, f, ensure_ascii=False, indent=1)
        f.write('\n')
    print(f'\nзаписано в {PLANTS}')


if __name__ == '__main__':
    main()
