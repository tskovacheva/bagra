#!/usr/bin/env python3
"""Merge the phase 2 workbook into seed/combinations.json and seed/sources.json.

Same rule as phase 1: FILLS ONLY. A cell that would change a recorded value is
held and printed, never applied.

Two things are different here and both are the point of the sheet.

`none` IS AN ANSWER. Twenty-two combination records carried `mordantCode: null`
and a code that says `nomordant` — the code claimed no mordant and the key had
never said so. Writing `none` turns „does not say" into „says: no mordant",
which is the distinction the reference engine was taught in §13ck. Blank stays
blank; only an explicit `none` in the sheet writes one.

THE CITATIONS ARE FREE TEXT and attribution belongs in the register (§13bt).
Each distinct citation is registered as a source and the record points at it by
code, rather than carrying a sentence nobody can look up.
"""
import json, re, sys, collections
from openpyxl import load_workbook

BOOK = '/mnt/user-data/uploads/Багра-фаза-2-комбинации-група-2.xlsx'

wb = load_workbook(BOOK)
combos_pack = json.load(open('seed/combinations.json'))
sources_pack = json.load(open('seed/sources.json'))
combos = {r['code']: r for r in combos_pack['combinations']}
known_sources = {s['code'] for s in sources_pack['sources']}

vocab = collections.defaultdict(set)
for m in re.finditer(r"V\('([a-z_]+)',\s*'([a-zA-Z0-9_]+)',", open('vocab.js').read()):
    vocab[m.group(1)].add(m.group(2))

filled, held, notes_added, skipped = [], [], 0, 0

# Each free-text citation becomes a code. Written out rather than derived from
# the string, so a code is a decision and not an accident of punctuation.
CITATIONS = {
    'Jenny Dean — Wild Colour': 'jenny-dean-wild-colour',
    'Jenny Dean — Wild Colour / Alkanet Root': 'jenny-dean-wild-colour',
    'George Weil — Oak Bark natural dye': 'george-weil-oak',
    'Eser et al., Journal of Natural Fibers 2017': 'eser-salvia-2017',
    'Natural Dyes for Textiles — Cosmos sulphureus': 'cosmos-sulphureus-textiles',
    'Safapour & Rather, Journal of the Textile Institute 2024': 'safapour-melissa-2024',
    'Molecules 2022 — pomegranate peel on cotton': 'molecules-peel-cotton-2022',
    'Molecules 2022 — onion & pomegranate peel on cotton': 'molecules-peel-cotton-2022',
    'Jenny Dean — Rhubarb Root / More Dye Extracts': 'jenny-dean-rhubarb',
    'Journal of the Textile Institute — eucalyptus leaf on cotton': 'eucalyptus-cotton-jti',
    'OpenLearn, Experiments with Natural Dyes': 'openlearn-natural-dyes',
    'Bhuiyan et al., Journal of Cleaner Production 2017': 'bhuiyan-henna-2017',
    'Rubia tinctorum on wool': 'rubia-wool-2022',
    'Lohar & Majumder 2019 — Tagetes erecta on cotton': 'lohar-tagetes-2019',
    'Ultrasound-assisted dyeing of cotton with marigold': 'ultrasound-marigold',
    'Farooq et al. 2013 / Tagetes erecta cotton': 'farooq-tagetes-2013',
    'Colorants 2025 — Persicaria tinctoria indigo on multifiber textiles': 'colorants-persicaria-2025',
}


def cite(text):
    """The code for a citation, and the URL it came with."""
    if not text:
        return None, None
    text = str(text).strip()
    url = None
    m = re.search(r'https?://\S+', text)
    if m:
        url = m.group(0).rstrip('.,;')
        text = text[:m.start()].rstrip(' ;,')
    for prefix, code in CITATIONS.items():
        if text.startswith(prefix):
            return code, url
    return None, url


def offer(where, field, current, incoming, apply):
    global skipped
    if incoming in (None, '', []):
        return
    if current in (None, '', []):
        apply()
        filled.append((where, field, incoming))
        return
    if str(current).strip() == str(incoming).strip():
        skipped += 1
        return
    held.append((where, field, current, incoming))


ws = wb['1 Влакно и мордант']
seen_citations = {}

for r in range(7, ws.max_row + 1):
    v = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    code = v[0]
    if not code or str(code).startswith('ПРИМЕР'):
        continue
    if not any(x not in (None, '') for x in v[6:11]):
        continue

    rec = combos.get(code)
    if not rec:
        held.append((code, 'record', 'not in the pack', '—'))
        continue

    key = rec['key']
    offer(code, 'fibreClass', key.get('fibreClass'), v[6],
          lambda k=key, x=v[6]: k.__setitem__('fibreClass', x))
    offer(code, 'mordantCode', key.get('mordantCode'), v[7],
          lambda k=key, x=v[7]: k.__setitem__('mordantCode', x))
    offer(code, 'mordantBand', key.get('mordantBand'), v[8],
          lambda k=key, x=v[8]: k.__setitem__('mordantBand', x))

    # „How does it differ" is prose about the record, and `notes` looked like
    # its home — until the merge found `notes` already occupied on all 22 rows,
    # and occupied by two different things. Some hold real prose („Орехът е
    # субстантивен — хваща без мордант"); twenty-two hold what are plainly
    # CONDITION LABELS („с железни соли", „кора, алуминиев мордант"), which look
    # like the `conditions` text from an earlier import landing in the wrong
    # field. `influences` is declared on every record and populated on none.
    #
    # So there are three candidate homes and no clear one. Held rather than
    # guessed: a sentence written by the owner filed under a heading nobody
    # chose is a sentence that will be found in a year and not trusted.
    if v[9]:
        cur = (rec.get('notes') or {}).get('bg')
        if not cur:
            rec.setdefault('notes', {'bg': '', 'en': ''})['bg'] = str(v[9]).strip()
            notes_added += 1
        else:
            held.append((code, 'notes', cur[:34], str(v[9])[:34]))

    scode, url = cite(v[10])
    if scode:
        seen_citations[scode] = (str(v[10]).strip(), url)
        # `learnedFrom` is the record's attribution. A code that resolves in the
        # register beats a sentence nobody can look up.
        offer(code, 'learnedFrom', rec.get('learnedFrom'), scode,
              lambda rr=rec, x=scode: rr.__setitem__('learnedFrom', x))
    elif v[10]:
        held.append((code, 'source', 'unrecognised citation', str(v[10])[:60]))

# ---------------------------------------------------------------- sources
NEW = {
    'eser-salvia-2017': ('reference', 'Dyeing of wool with Salvia officinalis and metal mordants',
        'Eser et al., Journal of Natural Fibers', None),
    'cosmos-sulphureus-textiles': ('reference', 'Natural Dyes for Textiles — Cosmos sulphureus', '—', None),
    'safapour-melissa-2024': ('reference', 'Dyeing of wool yarn with Melissa officinalis',
        'Safapour & Rather, Journal of the Textile Institute', None),
    'molecules-peel-cotton-2022': ('reference', 'Onion and pomegranate peel extracts on cotton',
        'Molecules', None),
    'jenny-dean-rhubarb': ('site', 'Rhubarb Root — More Dye Extracts', 'Jenny Dean', None),
    'eucalyptus-cotton-jti': ('reference', 'Eucalyptus leaf extract on cotton',
        'Journal of the Textile Institute', None),
    'openlearn-natural-dyes': ('course', 'Experiments with Natural Dyes', 'OpenLearn', None),
    'bhuiyan-henna-2017': ('reference', 'Henna on protein fibres',
        'Bhuiyan et al., Journal of Cleaner Production', None),
    'rubia-wool-2022': ('reference', 'Rubia tinctorum on wool with alum', '—', None),
    'lohar-tagetes-2019': ('reference', 'Tagetes erecta on cotton', 'Lohar & Majumder', None),
    'ultrasound-marigold': ('reference', 'Ultrasound-assisted dyeing of cotton with marigold', '—', None),
    'farooq-tagetes-2013': ('reference', 'Tagetes erecta on cotton', 'Farooq et al.', None),
    'colorants-persicaria-2025': ('reference', 'Persicaria tinctoria indigo on multifibre textiles',
        'Colorants', None),
}
added_sources = 0
for scode, (raw, url) in seen_citations.items():
    if scode in known_sources or scode not in NEW:
        continue
    kind, name, author, _ = NEW[scode]
    sources_pack['sources'].append({
        'code': scode, 'kind': kind, 'name': name, 'author': author or '—',
        'url': url or '',
        'note': {'bg': 'Регистриран при сливането на фаза 2. Цитиран за конкретна двойка '
                       'влакно и мордант, не за цвят.',
                 'en': 'Registered when phase 2 was merged. Cited for one fibre-and-mordant '
                       'pairing, never for colour.'},
    })
    known_sources.add(scode)
    added_sources += 1

if '--apply' in sys.argv:
    combos_pack['packVersion'] = '0.5.0'
    sources_pack['packVersion'] = 5
    json.dump(combos_pack, open('seed/combinations.json', 'w'), ensure_ascii=False, indent=1)
    json.dump(sources_pack, open('seed/sources.json', 'w'), ensure_ascii=False, indent=1)

print(f'FILLED    {len(filled)}   (empty → value)')
print(f'NOTES     {notes_added}   (a sentence about how it differs)')
print(f'SOURCES   {added_sources}   registered from free-text citations')
print(f'ALREADY   {skipped}')
print(f'HELD      {len(held)}   (would change a recorded value — NOT applied)')
print()
for k, n in collections.Counter(f for _, f, _ in filled).most_common():
    print(f'  {k:16} {n}')
print()
for h in held:
    print(f'  HELD  {h[0][:44]:44} {h[1]:12} {h[2]!r} → {h[3]!r}')
if '--apply' not in sys.argv:
    print('\n(dry run — pass --apply to write)')
