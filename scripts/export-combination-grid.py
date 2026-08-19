#!/usr/bin/env python3
"""Export a grid for writing combinations, seeded from the swatches already held.

The reference engine covers 10 plants of 57. The other 47 are not blank,
though: they carry 146 colour swatches between them, each with a hex, a name
and a free-text condition — „алуминиев мордант", „с желязо", „слаба баня". That
is a colour someone has already looked at and judged. It is not missing
information; it is information in the wrong place.

So this does not ask for colours to be invented. It lays each existing swatch
out as a draft row and asks for the ONE thing a swatch does not carry: the
structured key — which mordant, which pH, which process — that lets the
reference engine reach it.

The condition text is offered as a suggestion, never as an answer. Mapping
„алуминиев мордант" to `alum_potassium` is safe; „слаба баня" says nothing
about a mordant at all, and „лист с желязно одеало" is an eco print. A guess
written into the key would become a fact nobody checked, so the suggestion sits
in its own column and the key columns start empty where the text does not
plainly say.

Only cellulose, per the owner's call — all 28 existing records are cellulose,
and protein is a second pass, not a second column to leave half-filled.

Usage:  python3 scripts/export-combination-grid.py [output.xlsx]
"""
import json
import sys
import re
import collections

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PLANTS = 'seed/plants.json'
COMBOS = 'seed/combinations.json'
VOCAB = 'vocab.js'
DEFAULT_OUT = 'combination-grid.xlsx'

MORDANTS = ['none', 'alum_potassium', 'iron', 'copper', 'alum_acetate']
PROCESSES = ['immersion', 'ecoprint']
PH = ['', 'acid', 'neutral', 'alkaline']
CONFIDENCE = ['literature', 'practice', 'confirmed']

# Only what the words plainly say. Anything describing bath STRENGTH or
# DURATION is deliberately absent: it says nothing about the mordant, and the
# whole point of §13br was that strength is not a key dimension.
SUGGEST_MORDANT = [
    (r'алуминиев мордант|алуминиева стипца|със стипца', 'alum_potassium'),
    (r'железн|с желязо', 'iron'),
    (r'без мордант|без байц', 'none'),
    (r'медн|с мед', 'copper'),
    (r'алуминиев ацетат', 'alum_acetate'),
]
SUGGEST_PH = [
    (r'алкалн', 'alkaline'),
    (r'кисел', 'acid'),
]
SUGGEST_PROCESS = [
    (r'одеало|одеяло|върху мордантиран плат|отпечат', 'ecoprint'),
]


def vocab_pairs(kind):
    src = open(VOCAB, encoding='utf-8').read()
    out = {}
    for m in re.finditer(rf"'{kind}',\s*'([a-z_]+)',\s*'([^']*)',\s*'([^']*)'", src):
        if re.search(r'[а-яА-Я]', m.group(2)):
            out[m.group(1)] = m.group(2)
    return out


def suggest(text, table):
    for pattern, code in table:
        if re.search(pattern, text, re.I):
            return code
    return ''


def text_of(field, lang='bg'):
    if isinstance(field, dict):
        return field.get(lang) or field.get('en') or ''
    return field or ''


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT

    plants = json.load(open(PLANTS, encoding='utf-8'))['plants']
    combos = json.load(open(COMBOS, encoding='utf-8'))['combinations']
    part_names = vocab_pairs('plant_part')

    covered = {c['key']['dyeSource']['plantId'].replace('seed:', '') for c in combos}

    # What each covered plant already answers, so a row is not requested twice.
    have = set()
    for c in combos:
        k = c['key']
        have.add((k['dyeSource']['plantId'].replace('seed:', ''),
                  k['dyeSource'].get('partCode'),
                  k.get('mordantCode'),
                  (k.get('medium') or {}).get('phCode') if isinstance(k.get('medium'), dict) else None,
                  k.get('processCode')))

    rows = []
    for p in plants:
        pname = text_of(p.get('nameCommon'))
        parts = [pt.get('partCode') for pt in (p.get('parts') or [])]
        for sw in (p.get('colours') or []):
            cond = text_of(sw.get('conditions'))
            rows.append({
                'plant': pname,
                'code': p['code'],
                'part': part_names.get(parts[0], parts[0]) if len(parts) == 1 else '',
                'partHint': ' / '.join(part_names.get(x, x) for x in parts),
                'colour': text_of(sw.get('name')),
                'hex': sw.get('hex') or '',
                'cond': cond,
                'mordant': suggest(cond, SUGGEST_MORDANT),
                'ph': suggest(cond, SUGGEST_PH),
                'process': suggest(cond, SUGGEST_PROCESS) or 'immersion',
                'source': sw.get('source') or '',
                'confidence': sw.get('confidence') or '',
                'done': 'вече има' if p['code'] in covered else '',
            })

    rows.sort(key=lambda r: (r['done'] != '', r['plant'], r['cond']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Комбинации'

    headers = ['Растение', 'Код', 'Част', 'Възможни части', 'Цвят', 'HEX',
               'Условия (както са записани)', 'Байц', 'pH', 'Процес',
               'Сигурност', 'Състояние', 'Бележка']
    widths = [22, 22, 14, 22, 26, 10, 34, 18, 12, 14, 14, 12, 30]

    hf = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2C3B57')
    bf = Font(name='Arial', size=10)
    editfill = PatternFill('solid', fgColor='FFF3CD')
    donefont = Font(name='Arial', size=10, color='9A948A')
    thin = Side(style='thin', color='DED8CA')
    border = Border(bottom=thin)

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill = hf, hfill
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    for n, r in enumerate(rows, start=2):
        vals = [r['plant'], r['code'], r['part'], r['partHint'], r['colour'],
                r['hex'], r['cond'], r['mordant'], r['ph'], r['process'],
                r['confidence'], r['done'], '']
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=n, column=i, value=v)
            c.font = donefont if r['done'] else bf
            c.border = border
        if not r['done']:
            for col in (3, 8, 9, 10):
                ws.cell(row=n, column=col).fill = editfill
        if r['hex']:
            try:
                ws.cell(row=n, column=6).fill = PatternFill('solid', fgColor=r['hex'].lstrip('#'))
            except Exception:
                pass

    def dv(col, values, title):
        d = DataValidation(type='list', formula1='"' + ','.join(values) + '"',
                           allow_blank=True, showDropDown=False)
        d.errorTitle, d.error = title, 'Стойност извън речника'
        ws.add_data_validation(d)
        d.add(f'{get_column_letter(col)}2:{get_column_letter(col)}{len(rows) + 1}')

    dv(8, MORDANTS, 'Байц')
    dv(9, PH[1:], 'pH')
    dv(10, PROCESSES, 'Процес')
    dv(11, CONFIDENCE, 'Сигурност')

    # ---------------------------------------------------------------- legend
    ls = wb.create_sheet('Как се попълва')
    ls.column_dimensions['A'].width = 24
    ls.column_dimensions['B'].width = 84

    def line(row, a, b):
        ca, cb = ls.cell(row=row, column=1, value=a), ls.cell(row=row, column=2, value=b)
        ca.font = Font(name='Arial', size=10, bold=True)
        cb.font = Font(name='Arial', size=10)
        cb.alignment = Alignment(wrap_text=True, vertical='top')

    todo = sum(1 for r in rows if not r['done'])
    suggested = sum(1 for r in rows if not r['done'] and r['mordant'])

    line(1, 'Какво е това', 'Всяка цветна мостра в библиотеката, подредена като чернова за комбинация.')
    line(2, '', 'Цветът вече е твой — липсва само ключът, по който справочникът да го намери.')
    line(4, 'Какво се пипа', 'Жълтите клетки: Част, Байц, pH, Процес.')
    line(5, 'Байц', 'none · alum_potassium · iron · copper · alum_acetate')
    line(6, 'pH', 'празно (обикновена вода) · acid · alkaline')
    line(7, 'Процес', 'immersion · ecoprint')
    line(9, 'Редове общо', str(len(rows)))
    line(10, 'За попълване', f'{todo} — жълти, най-отгоре')
    line(11, 'С предложен байц', f'{suggested} — предложен от текста на условията; ПРОВЕРИ ГО')
    line(12, 'Сиви редове', 'Растения, които вече имат комбинации. За справка, не се пипат.')
    line(14, 'Предложенията', 'Само където текстът казва ясно. «слаба баня» не значи байц —')
    line(15, '', 'силата на банята не е измерение в ключа (§13br), а бележка на записа.')
    line(17, 'Ако не се знае', 'Остави празно. Редът просто няма да се слее — по-добре от грешен ключ.')
    line(19, 'Част', 'Попълнена автоматично само където растението има една част.')
    line(20, '', 'Където има няколко, виж «Възможни части» и избери.')

    wb.save(out_path)
    print(f'{out_path}: {len(rows)} реда, {todo} за попълване, {suggested} с предложен байц')


if __name__ == '__main__':
    main()
