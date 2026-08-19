#!/usr/bin/env python3
"""Merge the reviewed chemistry workbook back into seed/plants.json.

The counterpart to export-chemistry-audit.py. Reads the `Степен` column and
writes it onto the matching plant × part × compound.

What it will NOT do, deliberately:

- **Overwrite a level that is already set.** If the workbook disagrees with the
  library, it reports the disagreement and changes nothing. A bulk merge that
  silently reverses an earlier judgement turns a spreadsheet round-trip into an
  edit nobody reviewed.
- **Invent a chemistry entry.** A row for a part with no chemistry recorded is
  reported, not created: deciding a compound belongs on a part is a domain
  decision, and this script has no standing to make it.
- **Accept a value outside the vocabulary.** trace / moderate / high / dominant
  and nothing else. Anything else stops the run rather than entering the
  library as an unknown code.

Dry by default — prints what it would do. Pass --write to actually save.

Usage:
    python3 scripts/merge-chemistry-audit.py chemistry-audit.xlsx
    python3 scripts/merge-chemistry-audit.py chemistry-audit.xlsx --write
"""
import json
import sys
import re
import collections

from openpyxl import load_workbook

PLANTS = 'seed/plants.json'
VOCAB = 'vocab.js'
LEVELS = {'trace', 'moderate', 'high', 'dominant'}


def vocab_pairs(kind):
    src = open(VOCAB, encoding='utf-8').read()
    out = {}
    for m in re.finditer(rf"'{kind}',\s*'([a-z_]+)',\s*'([^']*)',\s*'([^']*)'", src):
        code, bg = m.group(1), m.group(2)
        if re.search(r'[а-яА-Я]', bg):
            out[bg] = code          # bg name -> code, the direction needed here
    return out


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    write = '--write' in sys.argv

    chem_code = vocab_pairs('chemistry_class')
    part_code = vocab_pairs('plant_part')

    wb = load_workbook(path, data_only=True)
    ws = wb['Химия']

    doc = json.load(open(PLANTS, encoding='utf-8'), object_pairs_hook=collections.OrderedDict)
    by_code = {p['code']: p for p in doc['plants']}

    applied, conflicts, skipped, unknown = [], [], [], []

    for row in ws.iter_rows(min_row=2, values_only=True):
        plant_bg, botanical, code, part_bg, compound_bg, level, status, note = (list(row) + [None] * 8)[:8]
        level = (level or '').strip()
        if not level:
            continue
        if level not in LEVELS:
            unknown.append((code, part_bg, compound_bg, level))
            continue
        if not compound_bg:
            skipped.append((code, part_bg, 'ред без вещество — не създавам записи'))
            continue

        plant = by_code.get(code)
        if not plant:
            skipped.append((code, part_bg, 'растението не е намерено'))
            continue

        pc = part_code.get(part_bg, part_bg)
        cc = chem_code.get(compound_bg, compound_bg)

        part = next((p for p in (plant.get('parts') or []) if p.get('partCode') == pc), None)
        if not part:
            skipped.append((code, part_bg, 'частта не е намерена'))
            continue

        entry = next((c for c in (part.get('chemistry') or []) if c.get('classCode') == cc), None)
        if not entry:
            skipped.append((code, part_bg, f'{compound_bg}: няма такъв запис — не създавам'))
            continue

        current = entry.get('level') or ''
        if current and current != level:
            conflicts.append((code, part_bg, compound_bg, current, level))
            continue
        if current == level:
            continue

        entry['level'] = level
        applied.append((code, part_bg, compound_bg, level))

    if unknown:
        print('НЕВАЛИДНИ СТОЙНОСТИ — нищо не е записано:')
        for u in unknown:
            print('  ', ' / '.join(str(x) for x in u))
        sys.exit(1)

    print(f'ще се запишат: {len(applied)}')
    for a in applied[:20]:
        print('  +', ' / '.join(str(x) for x in a))
    if len(applied) > 20:
        print(f'  ... и още {len(applied) - 20}')

    if conflicts:
        print(f'\nРАЗМИНАВАНИЯ — НЕ са пипнати ({len(conflicts)}):')
        for c in conflicts:
            print(f'  {c[0]} / {c[1]} / {c[2]}: в базата «{c[3]}», в файла «{c[4]}»')

    if skipped:
        print(f'\nпропуснати ({len(skipped)}):')
        for s in skipped[:20]:
            print('  -', ' / '.join(str(x) for x in s))
        if len(skipped) > 20:
            print(f'  ... и още {len(skipped) - 20}')

    if not write:
        print('\n(пробен пуск — нищо не е записано. Пусни с --write за да запишеш.)')
        return

    if applied:
        with open(PLANTS, 'w', encoding='utf-8') as f:
            json.dump(doc, f, ensure_ascii=False, indent=1)
            f.write('\n')
        print(f'\nзаписано в {PLANTS}')
    else:
        print('\nнищо за записване')


if __name__ == '__main__':
    main()
