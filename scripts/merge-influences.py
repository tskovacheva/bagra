#!/usr/bin/env python3
"""Bring in what the three data rounds returned and the model could not hold.

Thirty-nine explanatory texts came back across phases 2, 3 and 4 — „what makes
this result move" — each with a source, and each was HELD at import because a
combination could say what to expect and not why it changes. `notes` was already
occupied and by two different kinds of thing, `expected.variation` belongs to
the reference engine, and `influences` was declared on every record and
populated on none.

They are not new research. Every line below is read out of a workbook the owner
filled in; only the FACTOR is added here, and it is added by reading the
sentence, which is a judgement no regular expression makes.

Two model changes, both additive:

  sourceCodes  — a list, because a record can rest on more than one source and
                 they are not the same claim. `learnedFrom` stays exactly as it
                 is: the colour source of an existing record is not replaced by
                 an evidence source arriving beside it.
  influences   — [{factor, text:{bg,en}, sourceCode}], the field that has been
                 declared since the model was written and never filled.

Idempotent: run it twice and the second run reports nothing to do.
"""
import json, sys, collections
from openpyxl import load_workbook

BOOKS = [
    '/mnt/user-data/uploads/Багра-фаза-2-комбинации-група-2.xlsx',
    '/mnt/user-data/uploads/Багра-фаза-3-попълнено-сигурни.xlsx',
    '/mnt/user-data/uploads/Багра-фаза-4-попълнено-за-затваряне.xlsx',
]

combos_pack = json.load(open('seed/combinations.json'))
sources_pack = json.load(open('seed/sources.json'))
combos = {r['code']: r for r in combos_pack['combinations']}
known = {s['code'] for s in sources_pack['sources']}

# WHICH FACTOR EACH SENTENCE IS ABOUT, read by hand.
#
# „Желязото потъмнява тона" is about the mordant. „E. cinerea може да даде
# червено-оранжево, E. globulus по-кафяво" is about the species. „Памукът има
# по-нисък афинитет" is about the fibre. A classifier would have to understand
# the sentence, and there are thirty-nine of them.
FACTOR = {
    'alkanna_tinctoria_root_iron_immersion': 'mordant',
    'sambucus_nigra_bark_alum_potassium_immersion': 'fibre',
    'sambucus_nigra_fruit_nomordant_immersion': 'ph',
    'salvia_officinalis_leaf_alum_potassium_immersion': 'fibre',
    'salvia_officinalis_leaf_iron_immersion': 'mordant',
    'salvia_officinalis_leaf_nomordant_immersion': 'mordant',
    'eucalyptus_spp_leaf_iron_immersion': 'mordant',
    'eucalyptus_spp_leaf_nomordant_immersion': 'species',
    'cosmos_sulphureus_flower_alum_potassium_immersion': 'ph',
    'cosmos_sulphureus_flower_iron_immersion': 'mordant',
    'allium_cepa_hull_alum_potassium_immersion': None,   # source only, no text
    'allium_cepa_hull_nomordant_immersion': None,
    'lawsonia_inermis_leaf_nomordant_immersion': 'fibre',
    'melissa_officinalis_leaf_alum_potassium_immersion': 'mordant',
    'melissa_officinalis_leaf_iron_immersion': 'mordant',
    'punica_granatum_hull_alum_potassium_immersion': 'mordant',
    'punica_granatum_hull_nomordant_immersion': 'fibre',
    'punica_granatum_hull_iron_immersion': 'mordant',
    'rheum_rhabarbarum_root_alum_potassium_immersion': 'fibre',
    'rheum_rhabarbarum_root_nomordant_immersion': 'fibre',
    'rheum_rhabarbarum_root_nomordant_alkalinedye_bath_immersion': 'ph',
    'rubia_tinctorum_root_alum_potassium_immersion': 'concentration',
    'rubia_tinctorum_root_nomordant_immersion': 'mordant',
    'rubia_tinctorum_root_nomordant_alkalinedye_bath_immersion': 'ph',
    'quercus_robur_bark_nomordant_immersion': 'fibre',
    'tagetes_erecta_flower_alum_potassium_immersion': 'concentration',
    'tagetes_erecta_flower_nomordant_immersion': 'mordant',
    'tagetes_erecta_flower_iron_immersion': 'mordant',
    'persicaria_tinctoria_leaf_nomordant_immersion': 'fibre',
    'frangula_alnus_bark_nomordant_immersion': 'preparation',
    'frangula_alnus_bark_nomordant_alkalinedye_bath_immersion': 'ph',
    'rubus_fruticosus_leaf_alum_potassium_immersion': 'mordant',
    'rubus_fruticosus_leaf_iron_immersion': 'mordant',
    'calendula_officinalis_flower_alum_potassium_immersion': 'mordant',
    'calendula_officinalis_flower_iron_immersion': 'mordant',
    'rosa_spp_leaf_iron_immersion': 'species',
    'rosa_spp_leaf_nomordant_immersion': 'species',
    'betula_pendula_bark_nomordant_immersion': 'fibre',
    'mentha_spp_leaf_nomordant_immersion': 'fibre',
}

# The citation string in the workbook, to a code in the register. Written out so
# a code is a decision rather than an accident of punctuation.
CITE = [
    ('Jenny Dean — Wild Colour',              'jenny-dean-wild-colour'),
    ('Jenny Dean — Rhubarb Root',             'jenny-dean-rhubarb'),
    ('Jenny Dean — Anglo-Saxon',              'jenny-dean-anglosaxon'),
    ('Jenny Dean — Solar Dye Pots',           'jenny-dean-solar'),
    ('Jenny Dean — Focus on Tannin',          'jenny-dean-tannin'),
    ('Jenny Dean — Dyeing brown and grey',    'jenny-dean-brown-grey'),
    ('Eser et al.',                           'eser-salvia-2017'),
    ('Journal of Natural Fibers — Salvia',    'eser-salvia-2017'),
    ('Journal of the Textile Institute — eucalyptus', 'eucalyptus-cotton-jti'),
    ('Natural Dyes for Textiles',             'cosmos-sulphureus-textiles'),
    ('OpenLearn',                             'openlearn-natural-dyes'),
    ('Bhuiyan et al.',                        'bhuiyan-henna-2017'),
    ('Safapour & Rather',                     'safapour-melissa-2024'),
    ('Molecules 2022',                        'molecules-peel-cotton-2022'),
    ('Rubia tinctorum on wool',               'rubia-wool-2022'),
    ('George Weil',                           'george-weil-oak'),
    ('Lohar & Majumder',                      'lohar-tagetes-2019'),
    ('Ultrasound-assisted dyeing',            'ultrasound-marigold'),
    ('Farooq et al.',                         'farooq-tagetes-2013'),
    ('Colorants 2025',                        'colorants-persicaria-2025'),
    ('Catharine Ellis',                       'ellis-natural-dye'),
    ('Luhamaa et al.',                        'luhamaa-frangula-2025'),
    ('Cuce, Journal of Natural Fibers',       'cuce-rosa-canina'),
    ('Kuwait Journal of Science',             'kuwait-mentha-2025'),
]

# FREE TEXT THAT IS ALREADY A SOURCE. 105 records carry a sentence in
# `learnedFrom` where the register holds the same thing under a code — the guide
# most of all, on 102 of them. Left as prose, `sourceCodes` would point at
# strings that resolve to nothing, and the integrity guard would be checking a
# list of sentences.
#
# `learnedFrom` itself is NOT rewritten: it is what those records have always
# said, and rewriting it would be a migration performed to make a new field look
# tidy. The code goes into `sourceCodes`, which is the field that is meant to
# resolve.
PROSE_TO_CODE = [
    ('Ръководство НАТУРАЛНИ БАГРИЛА', 'crafty-place-guide'),
    ('Üren 2022',                     'avocado-fastness-2022'),
    ('Plants 2023',                   'plants-2023-origanum'),
    ('Koçak & Yılmaz 2025',           'kocak-rosmarinus-2025'),
]

NEW_SOURCES = {
    'plants-2023-origanum':  ('reference', 'Historical dye use of Origanum vulgare', 'Plants'),
    'kocak-rosmarinus-2025': ('reference', 'Rosmarinus officinalis leaves on wool',
                              'Koçak & Yılmaz'),
    'jenny-dean-anglosaxon': ('site', 'Anglo-Saxon Dye Experiments', 'Jenny Dean'),
    'jenny-dean-solar':      ('site', 'Solar Dye Pots', 'Jenny Dean'),
    'jenny-dean-tannin':     ('site', 'Focus on Tannin', 'Jenny Dean'),
    'jenny-dean-brown-grey': ('site', 'Dyeing brown and grey wool fibres', 'Jenny Dean'),
}


def as_code(value):
    """The register code for whatever `learnedFrom` holds — a code already, or
    the prose the register knows the same thing by."""
    if not value:
        return None
    if value in known:
        return value
    for prefix, code in PROSE_TO_CODE:
        if str(value).startswith(prefix):
            if code not in known and code in NEW_SOURCES:
                kind, name, author = NEW_SOURCES[code]
                sources_pack['sources'].append({
                    'code': code, 'kind': kind, 'name': name, 'author': author, 'url': '',
                    'note': {'bg': 'Регистриран при внасянето на обясненията.',
                             'en': 'Registered when the explanatory texts were imported.'}})
                known.add(code)
            return code if code in known else None
    return None


def cite(text):
    """(code, url) for a citation string, registering the source if it is new."""
    if not text:
        return None, None
    text = str(text).strip()
    url = None
    i = text.find('http')
    if i != -1:
        url = text[i:].split()[0].rstrip('.,;')
        text = text[:i].rstrip(' ;,')
    for prefix, code in CITE:
        if text.startswith(prefix):
            if code not in known:
                if code not in NEW_SOURCES:
                    return None, url          # unknown code, not invented
                kind, name, author = NEW_SOURCES[code]
                sources_pack['sources'].append({
                    'code': code, 'kind': kind, 'name': name, 'author': author,
                    'url': url or '',
                    'note': {'bg': 'Регистриран при внасянето на обясненията от работните '
                                   'книги. Цитиран за какво променя резултата, не за самия цвят.',
                             'en': 'Registered when the explanatory texts were imported from '
                                   'the workbooks. Cited for what moves the result, not for '
                                   'the colour itself.'}})
                known.add(code)
            return code, url
    return None, url


# ---------------------------------------------------------------- read
found = {}
for path in BOOKS:
    ws = load_workbook(path)['1 Влакно и мордант']
    for r in range(7, ws.max_row + 1):
        v = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        code = v[0]
        if not code or str(code).startswith('ПРИМЕР'):
            continue
        note, src = v[9], v[10]
        if not (note or src):
            continue
        # Later books win: phase 4 is the owner's latest reading of the same row.
        found[code] = (str(note or '').strip(), str(src or '').strip())

added_inf, added_src, held = 0, 0, []

# Every record, not only those with an explanation: `sourceCodes` is the field
# the integrity guard reads, and a record left without one would read as having
# no source at all when it has always had one.
resolved = 0
for rec in combos.values():
    if rec.get('sourceCodes') is not None:
        continue
    code = as_code(rec.get('learnedFrom'))
    if code:
        rec['sourceCodes'] = [code]
        resolved += 1
    elif rec.get('learnedFrom'):
        held.append((rec['code'], f'learnedFrom resolves to no source: {str(rec["learnedFrom"])[:40]}'))

for code, (note, src) in found.items():
    rec = combos.get(code)
    if not rec:
        held.append((code, 'no such combination record'))
        continue

    scode, url = cite(src)
    if src and not scode:
        held.append((code, f'unrecognised citation: {src[:44]}'))

    # ---- second source, alongside the first and not instead of it
    if scode:
        existing = rec.get('sourceCodes')
        if existing is None:
            # The colour source of an existing record stays where it is. A record
            # that already says „Ръководство НАТУРАЛНИ БАГРИЛА" learned its COLOUR
            # there; the paper arriving now taught which fibre and mordant. Both.
            existing = [x for x in [as_code(rec.get('learnedFrom'))] if x]
        if scode not in existing:
            existing.append(scode)
            added_src += 1
        rec['sourceCodes'] = existing

    # ---- the explanation
    if not note:
        continue
    factor = FACTOR.get(code)
    if not factor:
        held.append((code, 'no factor decided for this text'))
        continue
    inf = rec.setdefault('influences', [])
    if any(i.get('text', {}).get('bg') == note for i in inf):
        continue
    inf.append({'factor': factor, 'text': {'bg': note, 'en': ''},
                'sourceCode': scode or None})
    added_inf += 1

if '--apply' in sys.argv:
    combos_pack['packVersion'] = '0.8.0'
    sources_pack['packVersion'] = 9
    json.dump(combos_pack, open('seed/combinations.json', 'w'), ensure_ascii=False, indent=1)
    json.dump(sources_pack, open('seed/sources.json', 'w'), ensure_ascii=False, indent=1)

print(f'TEXTS READ    {len(found)}   across three workbooks')
print(f'INFLUENCES    {added_inf}   written to records')
print(f'SOURCES       {added_src}   second source references added')
print(f'RESOLVED      {resolved}   records whose single source became a code')
print(f'HELD          {len(held)}')
print()
for k, n in collections.Counter(
        i['factor'] for r in combos.values() for i in (r.get('influences') or [])).most_common():
    print(f'  {k:16} {n}')
print()
for c, why in held:
    print(f'  HELD  {c[:48]:48} {why}')
if '--apply' not in sys.argv:
    print('\n(dry run — pass --apply to write)')
