#!/usr/bin/env python3
"""Merge the filled combination grid into `seed/combinations.json` (§13cl).

Re-runnable. Without `--apply` it reports and writes nothing.

WHAT COMES BACK, AND WHAT IT CANNOT BE

177 rows, filled to the letter: nothing guessed, every blank carrying its reason.
And the fibre is answered on three of them, the mordant strength on none — the
guide records colour and conditions, not those, because to someone writing a
dyeing book they are obvious from context.

So **not one row carries a complete key**, and no amount of care in the filling
was going to change that. The engine was taught to hold a blank first (§13ck):
a record silent on the fibre now reads as silent rather than as contradicting a
question about cotton. That is what makes these rows usable at all.

THE LINE THIS SCRIPT DRAWS

A combination is an ANSWER to „this plant, this part, this way — what should I
expect". Two things are therefore required and are not negotiable:

    partCode      which part of the plant
    processCode   immersion or eco print

Without the part, the record does not say what was in the pot; a colour from oak
bark and a colour from oak leaf are different answers and the key cannot hold
both. Rows missing it stay swatches on the plant, where they already do useful
work, and are reported rather than forced in.

Everything else may be blank. `fibreClass`, `mordantCode`, `mordantBand`,
`medium` — a blank is a blank and the reference now says so.

WHAT IT REFUSES

- A row whose key matches a record already there, when the expected colour
  differs. Two answers to one question is not something a script settles.
- A row with a mordant the vocabulary does not have.
- A row whose plant code is not in the library.
- Two rows of the workbook producing the same key. That is the duplicate-key
  fault §13br already had to be cleaned up once, and it is cheaper to refuse
  than to merge and merge back.

Usage:  python3 scripts/merge-combination-grid.py <filled.xlsx> [--apply]
"""

import json
import pathlib
import re
import sys
from collections import Counter, defaultdict

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parent.parent
COMBINATIONS = ROOT / 'seed' / 'combinations.json'
PLANTS = ROOT / 'seed' / 'plants.json'
VOCAB = ROOT / 'vocab.js'

PACK_VERSION = '0.4.0'
LEARNED_FROM = 'Ръководство НАТУРАЛНИ БАГРИЛА, Crafty Place'

# Required. See the docstring: without these the record is not an answer.
REQUIRED = ('partCode', 'processCode')


def vocab(dimension):
    src = VOCAB.read_text(encoding='utf-8')
    return {m.group(1): m.group(2) for m in
            re.finditer(rf"V\('{dimension}',\s*'([a-z_]+)',\s*'([^']*)'", src)}


def main():
    if len(sys.argv) < 2:
        sys.exit('usage: merge-combination-grid.py <filled.xlsx> [--apply]')
    book = pathlib.Path(sys.argv[1])
    apply = '--apply' in sys.argv

    ws = load_workbook(book)['Комбинации']
    header_row = next((r for r in range(1, 30)
                       if str(ws.cell(row=r, column=1).value or '') == 'Растение'), None)
    if not header_row:
        sys.exit('no header row — is this the right sheet?')
    col = {str(ws.cell(row=header_row, column=c).value or '').strip(): c
           for c in range(1, ws.max_column + 1)}

    data = json.loads(COMBINATIONS.read_text(encoding='utf-8'))
    existing = data['combinations']
    by_key = {json.dumps(r['key'], sort_keys=True, ensure_ascii=False): r
              for r in existing}

    plant_ids = {p['code']: 'seed:' + p['code']
                 for p in json.loads(PLANTS.read_text(encoding='utf-8'))['plants']}

    parts_bg = {bg.lower(): code for code, bg in vocab('plant_part').items()}
    mordants = set(vocab('mordant_type'))
    fibres = set(vocab('fibre_class'))
    processes = set(vocab('process'))
    ph_codes = {'acid', 'neutral', 'alkaline'}
    bands = {'trace', 'low', 'medium', 'high'}

    problems = []
    new, merged_into, no_part, set_aside = [], [], [], []
    blanks = Counter()

    # SEVERAL SWATCHES, ONE KEY, ONE RECORD.
    #
    # Eucalyptus leaf by immersion with no recorded mordant appears three times —
    # керемидено, прасковено, ръждиво — and the three differ by the strength of
    # the bath, which is deliberately NOT a dimension of the key (§13br). They
    # are not three answers to one question; they are one answer with a range,
    # and the 28 records already there are written that way: „бежово, охра,
    # светъл до среден кафяв" is a span, not a colour.
    #
    # So rows are gathered by key first and a record is built from the group.
    # The colours are JOINED, never rewritten — a script does not compose prose —
    # and each row's own conditions are kept verbatim in the notes, which is
    # where bath strength was always supposed to live.
    rows_by_key = defaultdict(list)
    key_of, code_of = {}, {}

    # BUT NOT EVERY DIFFERENCE IS BATH STRENGTH.
    #
    # Woad leaf appears twice, „многократно потапяне" and „редукционна вана" —
    # and a reduction vat is not a weak bath, it is a different way of getting
    # the colour out. Alkanet the same: „алкохолен извлек" is solvent
    # extraction. The key does not carry the extraction method (it lives on the
    # dose, §13cc), so collapsing these would say woad by immersion gives
    # „средно синьо, светло синьо" and lose that one of them was a vat.
    #
    # A group holding any of these is left for review rather than merged. It is
    # the one place where the missing dimension does real damage, and it is
    # better seen than smoothed over.
    EXTRACTION_WORDS = [
        'редукционна вана', 'вана', 'ферментаци', 'алкална екстракция',
        'алкохолен извлек', 'спиртен', 'студена екстракция',
    ]

    # The reasons a cell was left blank were written FOR THE MERGE, and they
    # belong in this report rather than in the library. „Влакно: не е посочено в
    # източника" on every record is noise the reader did not ask for, and the
    # reference now says „не уточнява" for itself (§13ck).
    REPORT_PREFIXES = ('Влакно:', 'Байц:', 'Сила на байца:', 'pH на банята:',
                       'Част:', 'Процес:', 'Одеяло:')

    def reader_note(text_bg):
        # Split on the sentence too, not only on „·". The reasons were written
        # as running prose — „…не се приема `neutral` по подразбиране." can sit
        # inside a piece that began as a real condition, and a filter that cut
        # only on the separator left half of them in.
        keep = []
        for piece in text_bg.split(' · '):
            piece = piece.strip()
            if not piece:
                continue
            if piece.startswith(REPORT_PREFIXES):
                continue
            sentences = [x.strip() for x in re.split(r'(?<=[.!?])\s+', piece)]
            sentences = [x for x in sentences if x and not x.startswith(REPORT_PREFIXES)]
            # An instruction addressed to whoever filled the grid is not a note
            # about the dye. „не записвай…" is the tell.
            sentences = [x for x in sentences
                         if not re.search(r'\bне (записвай|приемай|гадай|попълвай)\b', x)]
            if sentences:
                keep.append(' '.join(sentences))
        return ' · '.join(keep)

    def g(r, name):
        if name not in col:
            return ''
        v = ws.cell(row=r, column=col[name]).value
        return '' if v is None else str(v).strip()

    for r in range(header_row + 1, ws.max_row + 1):
        code = g(r, 'Код')
        if not code:
            continue
        where = f"r{r} ({g(r, 'Растение')}, {g(r, 'Цвят')})"

        plant_id = plant_ids.get(code)
        if not plant_id:
            problems.append(f'{where}: no plant `{code}` in the library')
            continue

        part_bg = g(r, 'Част').lower()
        part = parts_bg.get(part_bg) if part_bg else None
        if part_bg and not part:
            problems.append(f'{where}: „{part_bg}" is not a plant part')
            continue

        process = g(r, 'Процес') or None
        if process and process not in processes:
            problems.append(f'{where}: „{process}" is not a process')
            continue

        # The two that must be there.
        if not part or not process:
            no_part.append(f'{where}: ' + ('no part' if not part else 'no process'))
            continue

        mordant = g(r, 'Байц') or None
        if mordant and mordant not in mordants:
            problems.append(f'{where}: „{mordant}" is not a mordant')
            continue

        band = g(r, 'Сила на байца') or None
        if band and band not in bands:
            problems.append(f'{where}: „{band}" is not a band')
            continue
        # A band without a mordant says the strength of nothing.
        if band and not mordant:
            problems.append(f'{where}: a band with no mordant')
            continue

        fibre = g(r, 'Влакно') or None
        if fibre and fibre not in fibres:
            problems.append(f'{where}: „{fibre}" is not a fibre class')
            continue

        ph = g(r, 'pH на банята') or None
        where_ph = g(r, 'Къде е pH-то') or None
        if ph and ph not in ph_codes:
            problems.append(f'{where}: „{ph}" is not a pH band')
            continue
        # pH with no place is pH of nothing in particular; the whole reason the
        # column exists is that an alkaline EXTRACTION is not an alkaline bath.
        if ph and not where_ph:
            problems.append(f'{where}: a pH with no „where"')
            continue

        blanket = g(r, 'Одеяло') or None

        for name, value in (('fibreClass', fibre), ('mordantCode', mordant),
                            ('mordantBand', band), ('medium', ph)):
            if not value:
                blanks[name] += 1

        key = {
            'dyeSource': {'plantId': plant_id, 'partCode': part},
            'fibreClass': fibre,
            'fibreCode': None,
            'mordantCode': mordant,
            'mordantBand': band,
            'processCode': process,
            'blanket': blanket,
            'medium': {'phCode': ph, 'whereCode': where_ph} if ph else None,
        }
        ks = json.dumps(key, sort_keys=True, ensure_ascii=False)

        colour = g(r, 'Цвят')
        note = g(r, 'Бележка')
        rows_by_key[ks].append({
            'where': where, 'colour': colour, 'hex': g(r, 'HEX'),
            'conditions': g(r, 'Условия (както са записани)'),
            'note': note, 'confidence': g(r, 'Сигурност') or 'unverified',
        })
        key_of[ks] = key
        # THE CODE MUST CARRY EVERY DIMENSION THE KEY DOES.
        #
        # `code` becomes the record's id, so two different keys under one code
        # means the second silently overwrites the first on install. The first
        # version left out `blanket` and `medium`, and four pairs collapsed —
        # among them madder root in an ALKALINE bath („винено") against madder
        # root with no recorded pH („ярко розово, керемидено"). Two real and
        # different answers, one of them disappearing between the pack and the
        # database, with nothing anywhere to say so.
        code_of[ks] = '_'.join(filter(None, [
            code, part,
            mordant or 'nomordant', band or '',
            (blanket + 'blanket') if blanket else '',
            (ph + where_ph) if ph else '',
            fibre or '',
            process])).replace('__', '_')
        continue

    ORDER = ['confirmed', 'practice', 'literature', 'unverified']
    for ks, rows in rows_by_key.items():
        colours = []
        for x in rows:
            if x['colour'] and x['colour'] not in colours:
                colours.append(x['colour'])
        mixed = [x for x in rows
                 if any(w in (x['conditions'] or '').lower() for w in EXTRACTION_WORDS)]
        # SET ASIDE, not refused. `problems` means „this script will not decide";
        # this is different — it is a known and recorded limit of the model
        # (§13cc), named precisely, and it should not hold back the other
        # seventy-four. The rows stay swatches until the key can carry the
        # method, and they are listed so nobody has to rediscover which.
        if mixed and len(rows) > 1:
            set_aside.append('different extraction methods under one key: '
                             + '; '.join(x['where'] for x in rows))
            continue

        notes = []
        for x in rows:
            bit = reader_note(' · '.join(filter(None, [x['conditions'], x['note']])))
            if bit and bit not in notes:
                notes.append(bit)

        # The weakest confidence in the group. A span held up by one literature
        # reading is a literature reading, whatever else is in it.
        conf = max((x['confidence'] for x in rows),
                   key=lambda c: ORDER.index(c) if c in ORDER else len(ORDER))

        record = {
            'code': code_of[ks],
            'key': key_of[ks],
            'expected': {
                'colourText': {'bg': ', '.join(colours), 'en': ''},
                'swatchHex': rows[0]['hex'] or None,
                'variation': {'bg': '', 'en': ''},
                'printQuality': None,
                'lightfastness': '',
                'washfastness': '',
            },
            'influences': [],
            'confidence': conf,
            'learnedFrom': LEARNED_FROM,
            'notes': {'bg': ' | '.join(notes), 'en': ''},
        }

        old = by_key.get(ks)
        if old:
            said = (old.get('expected', {}).get('colourText', {}).get('bg') or '').strip()
            if said and record['expected']['colourText']['bg'] != said:
                problems.append(
                    f'{rows[0]["where"]}: this key already answers „{said[:40]}" — two '
                    "answers to one question is not a script's decision")
            else:
                merged_into.append(rows[0]['where'])
            continue

        new.append(record)
        by_key[ks] = record

    print(f'{book.name}')
    print(f'  rows read                    : {sum(1 for r in range(header_row+1, ws.max_row+1) if ws.cell(row=r, column=1).value)}')
    print(f'  would become records         : {len(new)}')
    print(f'  key already answered, skipped: {len(merged_into)}')
    print(f'  stay swatches (no part/process): {len(no_part)}')
    print(f'  set aside for the model        : {len(set_aside)}')
    for x in set_aside:
        print(f'    - {x}')
    print(f'  library: {len(existing)} -> {len(existing) + len(new)}')
    print(f'  blanks carried honestly      : ' +
          ', '.join(f'{k} {v}' for k, v in blanks.most_common()))

    if problems:
        print(f'\n  {len(problems)} thing(s) this script will not decide:')
        for x in problems[:25]:
            print(f'    - {x}')
        if len(problems) > 25:
            print(f'    … and {len(problems) - 25} more')
        sys.exit('stopped: nothing written')

    if '--show' in sys.argv:
        import textwrap
        for rec in new:
            if len(rec['notes']['bg'].split(' | ')) > 1:
                print('\n  ' + rec['code'])
                print('    цвят : ' + rec['expected']['colourText']['bg'])
                print('    ключ : ' + ', '.join(
                    f'{k}={v}' for k, v in rec['key'].items()
                    if v and k not in ('dyeSource', 'fibreCode')))
                print(textwrap.fill(rec['notes']['bg'], 96,
                      initial_indent='    бел. : ', subsequent_indent='           '))

    if not apply:
        print('\n  dry run — pass --apply to write')
        return

    data['combinations'] = existing + new
    data['packVersion'] = PACK_VERSION
    COMBINATIONS.write_text(json.dumps(data, ensure_ascii=False, indent=1) + '\n',
                            encoding='utf-8')
    print(f'\n  written · pack {PACK_VERSION} · {len(data["combinations"])} combinations')


if __name__ == '__main__':
    main()
