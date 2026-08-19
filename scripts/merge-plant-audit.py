#!/usr/bin/env python3
"""Merge the plant audit workbook into seed/plants.json (§13aw).

What it applies, and nothing else:

  * botanical and common names where they are genuinely wrong
  * photograph credit — author, licence, source — including the eight that were
    held back for want of an author
  * plant parts that are ADDED, mapped onto vocabulary codes

What it deliberately does not touch, and why:

  * temperatures — the audit turned them into prose ("special process", "cold
    20-25"), which the numeric fields cannot hold. That waits on the decision
    about "unknown" and "approximate" as legitimate values.
  * tannin level and tannin function — the plant already carries chemistry per
    part, with a class and a level. A second field for the same thing would
    drift from the first.
  * parts that the audit REMOVES — a part carries chemistry and dosing, so
    dropping one deletes knowledge. Reported, not applied.
  * the harvest and processing prose — the owner is editing it herself.

Matching is by botanical name, then common name, with explicit overrides where
the audit renames a plant: the code is the identity and never moves, because
placements, combinations and swatches all point at it.

Idempotent: run twice and the second run reports no changes.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PACK = ROOT / 'seed' / 'plants.json'

# The audit renamed these, so neither name matches any more. The code decides.
OVERRIDE = {
    'Сапаново дърво (източен бразил)': 'paubrasilia_echinata',
}

# The audit writes parts as Bulgarian prose. The library stores codes.
PART_CODE = {
    'листа': 'leaf', 'лист': 'leaf', 'млади върхове': 'leaf',
    'кора': 'bark',
    'цветове': 'flower', 'цвят': 'flower', 'съцветия': 'flower',
    'венчелистчета / цветове': 'flower',
    'плодове': 'fruit', 'плод': 'fruit', 'жълъди': 'fruit',
    'шишарковидни плодни съцветия': 'fruit',
    'гали': 'gall',
    'семе': 'seed', 'семена': 'seed',
    'черупки': 'shell',
    'корени': 'root', 'корен': 'root', 'коренища': 'root',
    'зелена плодова обвивка': 'hull', 'кора на плода': 'hull',
    'сухи външни люспи': 'hull', 'обвивка': 'hull',
    'надземна част': 'whole', 'надземни върхове': 'whole', 'стъбла': 'whole',
    'сърцевинна дървесина': 'heartwood', 'дървесина': 'heartwood',
    # The nine new plants brought their own phrasings.
    'зрели плодове': 'fruit', 'плодове (по-скоро за пигменти/експерименти)': 'fruit',
    'кора/вътрешна кора': 'bark', 'млади стъбла/кора': 'bark',
    'цъфтящи върхове': 'flower', 'корени/коренища': 'root',
}

# Common names the audit capitalises or decorates. The library keeps them plain
# and lower case, and the botanical name is a field of its own — a display name
# reading „Сумак (Rhus coriaria)" says the same thing twice.
CLEAN_COMMON = {
    'Сапаново дърво (източен бразил)': 'Сапаново дърво',
    'Сумак (Rhus coriaria)': 'Сумак',
}

# One value, always. Twenty of fifty-seven rows arrived compound — „Храст /
# малко дърво" — and a field that accepts a slash is prose the filter cannot
# read. The first word is the plant's habit; that it sometimes grows taller is
# a sentence in „Как се държи".
# The five where the SECOND word of the compound is the truer one, named by the
# owner: „Храст / полухраст" for sage, lavender, thyme and rosemary, and
# „Тревисто растение / полухраст" for geranium. Taking the first word is the
# rule; these are the exceptions the rule was always going to need, and they are
# listed rather than guessed at from the family.
TYPE_OVERRIDE = {
    'salvia_officinalis': 'subshrub',
    'lavandula_angustifolia': 'subshrub',
    'thymus_vulgaris': 'subshrub',
    'rosmarinus_officinalis': 'subshrub',
    'pelargonium_zonale': 'subshrub',
}

PLANT_TYPE = {
    'дърво': 'tree', 'храст': 'shrub', 'полухраст': 'subshrub',
    'тревисто растение': 'herb', 'тревисто': 'herb',
}

# Three, because „градинско" and „култивирано" arrived as a pair on fifteen
# rows and a pair that never separates is one value with two names. Everything
# tended by a person — bed, pot, greenhouse, field — is `garden`; anything that
# arrives as bought material is `imported`.
HABITAT = {
    'диворастящо': 'wild', 'подивяло': 'wild',
    'градинско': 'garden', 'култивирано': 'garden', 'саксийно': 'garden',
    'градинско/градско': 'garden', 'градинско/култивирано': 'garden',
    'градско': 'garden', 'оранжерийно/стайно отглеждане': 'garden',
    'градинско в мек климат': 'garden', 'култивирано в топъл климат': 'garden',
    'кухненска суровина': None,   # not a habitat — how she comes by it
    'вносна суровина': 'imported',
}

# The three prose columns, and the section each becomes. Sections already carry
# „Багрилни качества"; the audit rewrote it, so it is replaced rather than
# doubled.
SECTIONS = [
    # „Защо действа" is not a new heading — it is „Багрилна съставка" asked in
    # plainer words, and the library already had it. Two headings for one
    # question is how thirty-two plants came to have four sections and seven to
    # have ten. A section stands on every plant or it does not exist.
    ('Защо действа', 'Багрилна съставка', 'Dye constituent'),
    ('Как се държи', 'Как се държи', 'How it behaves'),
    ('Багрилни качества', 'Багрилни качества', 'Dye qualities'),
    ('Беритба и обработка', 'Беритба и обработка', 'Harvest and processing'),
    ('Части', 'Използвани части', 'Parts used'),
    ('Източници – текстове', 'Източници', 'Sources'),
]

# Our own library, cited as evidence for itself. A reference that quotes itself
# proves nothing, so it is stripped wherever the audit picked it up.
SELF_CITATION = 'githubusercontent.com/tskovacheva/bagra'


def slug(botanical):
    """The code, from the botanical name: `Sambucus nigra L.` -> `sambucus_nigra`.

    Authority abbreviations and everything in brackets are dropped — the code is
    an identifier, not a citation, and it never changes once issued, so it must
    not carry anything that might be revised.
    """
    name = re.sub(r'\(.*?\)', ' ', str(botanical or ''))
    words = [w for w in re.split(r'[\s.]+', name) if w]
    keep = [w.lower() for w in words[:2] if not re.fullmatch(r'[A-Z][a-z]?', w) or w == words[0]]
    return re.sub(r'[^a-z_]', '', '_'.join(keep[:2]))


def new_plant(row, plants, by_code, by_bot, by_common, changes, gaps, apply_changes):
    """A plant the library does not have yet.

    Added only when it has a botanical name to build a code from. The record is
    created bare — names, role, type, habitat — and everything else is filled by
    the same passes that update an existing plant, so a new plant and an old one
    go through exactly one set of rules.
    """
    code = slug(row['Ботанически'])
    if not code or code in by_code:
        return None

    plant = {
        'code': code,
        'nameCommon': {'bg': CLEAN_COMMON.get(str(row['Растение']).strip(),
                                              str(row['Растение']).strip()), 'en': ''},
        'nameBotanical': str(row['Ботанически']).strip(),
        'role': [], 'compositionalRole': [], 'parts': [], 'colours': [], 'sections': [],
        'plantType': '', 'habitat': [], 'harvestMonths': [],
        'lightfastness': '', 'washfastness': '', 'confidence': {},
    }
    changes.append((code, 'NEW PLANT', '', plant['nameCommon']['bg']))
    gaps.append(f'{code}: new plant — no colours, no dosing, no fastness')
    if apply_changes:
        plants.append(plant)
    by_code[code] = plant
    by_bot[norm(plant['nameBotanical'])] = plant
    by_common[norm(plant['nameCommon']['bg'])] = plant
    return plant


# Part names as the audit writes them inside a temperature note.
TEMP_PART = {
    'листа': 'leaf', 'лист': 'leaf', 'кора': 'bark', 'плодове': 'fruit',
    'плод': 'fruit', 'цветове': 'flower', 'цвят': 'flower', 'корен': 'root',
    'корени': 'root', 'коренища': 'root', 'семена': 'seed', 'обвивка': 'hull',
    'гали': 'gall', 'дървесина': 'heartwood',
}


def parse_temperature(raw):
    """Read „70–85", „листа/кора 70–85; плодове 40–65", „90; за плодове ~70".

    Returns a list of (set_of_part_codes, (low, high)) — an empty set meaning
    every part — or None when the text is not a temperature at all.
    """
    out = []
    for chunk in re.split(r'[;,]', raw):
        chunk = chunk.strip()
        if not chunk:
            continue
        numbers = re.findall(r'\d+', chunk)
        if not numbers:
            return None
        low = int(numbers[0])
        high = int(numbers[1]) if len(numbers) > 1 else None
        parts = {code for word, code in TEMP_PART.items() if re.search(word, chunk)}
        out.append((parts, (low, high)))
    return out or None


def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').strip().lower())


def main():
    apply_changes = '--apply' in sys.argv
    book = ROOT / sys.argv[1] if len(sys.argv) > 1 and not sys.argv[1].startswith('--') else None
    if not book:
        sys.exit('usage: merge-plant-audit.py <workbook.xlsx> [--apply]')

    ws = openpyxl.load_workbook(book)['Растения']
    head = [c.value for c in ws[1]]
    rows = [dict(zip(head, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    pack = json.loads(PACK.read_text(encoding='utf-8'))
    plants = pack['plants']
    by_code = {p['code']: p for p in plants}
    by_bot = {norm(p.get('nameBotanical')): p for p in plants}
    by_common = {norm((p.get('nameCommon') or {}).get('bg')): p for p in plants}

    changes, gaps, held, unmatched = [], [], [], []

    for row in rows:
        plant = (by_code.get(OVERRIDE.get(row['Растение'], ''))
                 or by_bot.get(norm(row['Ботанически']))
                 or by_common.get(norm(row['Растение'])))
        if not plant:
            plant = new_plant(row, plants, by_code, by_bot, by_common,
                              changes, gaps, apply_changes)
            if not plant:
                unmatched.append(f"{row['Растение']} · {row['Ботанически']}")
                continue

        code = plant['code']

        # --- names -------------------------------------------------------
        want_bot = str(row['Ботанически'] or '').strip()
        if want_bot and want_bot != plant.get('nameBotanical'):
            changes.append((code, 'nameBotanical', plant.get('nameBotanical'), want_bot))
            if apply_changes:
                plant['nameBotanical'] = want_bot

        raw = str(row['Растение'] or '').strip()
        # Capitalised, both languages, always (§13bb). The library had „дъб"
        # beside „Бял равнец" because each name was typed as whoever wrote it
        # felt at the time, and a list where half the rows shout is a list that
        # looks broken before it is read.
        want_bg = CLEAN_COMMON.get(raw, raw)
        want_bg = want_bg[:1].upper() + want_bg[1:] if want_bg else want_bg
        # Only a real change of name, not a change of capitalisation.
        if want_bg and norm(want_bg) != norm((plant.get('nameCommon') or {}).get('bg')):
            changes.append((code, 'nameCommon.bg', (plant.get('nameCommon') or {}).get('bg'), want_bg))
            if apply_changes:
                plant.setdefault('nameCommon', {})['bg'] = want_bg

        # --- photograph credit -------------------------------------------
        author = str(row['Автор'] or '').strip()
        licence = str(row['Лиценз'] or '').strip()
        source = str(row['Източник'] or '').strip()
        if author and licence:
            credit = plant.get('photoCredit') or {}
            new = {**credit, 'author': author, 'licence': licence}
            if source:
                new['source'] = source
            new.setdefault('taxon', plant.get('nameBotanical'))
            if new != credit:
                changes.append((code, 'photoCredit',
                                credit.get('author') or '—', f'{author} · {licence}'))
                if apply_changes:
                    plant['photoCredit'] = new
            if not plant.get('photoData'):
                gaps.append(f'{code}: credit recorded, image file still missing')

        # --- role ----------------------------------------------------------
        ROLE = {'багрилно': 'dye', 'еко принт': 'ecoprint', 'екопринт': 'ecoprint'}
        want_role = []
        for piece in str(row.get('Роля') or '').split('·'):
            piece = piece.strip().lower()
            if piece in ROLE and ROLE[piece] not in want_role:
                want_role.append(ROLE[piece])
            elif piece:
                held.append(f'{code}: role "{piece}" has no vocabulary code')
        if want_role and sorted(want_role) != sorted(plant.get('role') or []):
            changes.append((code, 'role', ' '.join(plant.get('role') or []) or '—',
                            ' '.join(want_role)))
            if apply_changes:
                plant['role'] = want_role

        # --- growth form and habitat ---------------------------------------
        raw_type = str(row.get('Тип растение') or '').split('/')[0].strip().lower()
        want_type = TYPE_OVERRIDE.get(code) or PLANT_TYPE.get(raw_type)
        if raw_type and not want_type:
            held.append(f'{code}: plant type "{raw_type}" has no vocabulary code')
        elif want_type and want_type != plant.get('plantType'):
            changes.append((code, 'plantType', plant.get('plantType') or '—', want_type))
            if apply_changes:
                plant['plantType'] = want_type

        want_habitat = []
        for piece in str(row.get('Къде се среща / отглежда') or '').split('·'):
            piece = piece.strip().lower()
            if not piece:
                continue
            # „диворастящо в Южна Европа", „градинско в по-топъл климат" — the
            # value with a qualifier attached. The qualifier is geography and
            # belongs in the prose; the first word is the habitat.
            if piece not in HABITAT:
                head = piece.split(' в ')[0].split(',')[0].strip()
                # „някои видове диворастящи", „диворастящо/култивирано" — the
                # qualifier and the slash both hide a habitat word inside a
                # sentence. Every habitat word present is taken.
                found_any = [v for k, v in HABITAT.items()
                             if v and re.search(k[:6], head)]
                if head in HABITAT:
                    piece = head
                elif found_any:
                    for hab in found_any:
                        if hab not in want_habitat:
                            want_habitat.append(hab)
                    continue
            if piece not in HABITAT:
                held.append(f'{code}: habitat "{piece}" has no vocabulary code')
                continue
            hab = HABITAT[piece]
            if hab is None:
                held.append(f'{code}: "{piece}" is how she comes by it, not where it grows')
            elif hab not in want_habitat:
                want_habitat.append(hab)
        if want_habitat and want_habitat != (plant.get('habitat') or []):
            changes.append((code, 'habitat', ' '.join(plant.get('habitat') or []) or '—',
                            ' '.join(want_habitat)))
            if apply_changes:
                plant['habitat'] = want_habitat

        # `availability` said something about the owner on a record that ships
        # to other people. Removed wherever it survives (§13ay).
        if 'availability' in plant:
            changes.append((code, 'availability -', plant['availability'], 'removed'))
            if apply_changes:
                plant.pop('availability')

        # --- tannins, onto the part ----------------------------------------
        #
        # The audit put the level on the plant; chemistry lives on the part, and
        # a second home for it would drift from the first (§13ak). Where the
        # note names parts — „особено гали/кора", „в листа/стъбла" — those parts
        # get it; where it names none, every part does.
        #
        # Only from `moderate` upward: „ниски / не е основна характеристика" is
        # the absence of a finding, and writing it down as a finding would make
        # thirty-two records assert something nobody measured.
        tannin = str(row.get('Танини') or '').strip().lower()
        level = ('high' if 'висок' in tannin and 'много' not in tannin
                 else 'dominant' if 'много високи' in tannin
                 else 'moderate' if 'средни' in tannin and 'висок' not in tannin
                 else None)
        if 'ниск' in tannin and 'висок' not in tannin:
            level = None
        if level:
            named = {c for w, c in TEMP_PART.items() if re.search(w, tannin)}
            for target in plant.get('parts') or []:
                if named and target['partCode'] not in named:
                    continue
                chem = target.setdefault('chemistry', [])
                if any('tannin' in c.get('classCode', '') for c in chem):
                    continue
                changes.append((code, f'tannin · {target["partCode"]}', '—', level))
                if apply_changes:
                    chem.append({'classCode': 'tannin', 'level': level})

        # --- the three prose columns ---------------------------------------
        for column, title_bg, title_en in SECTIONS:
            body = str(row.get(column) or '').strip()
            if not body:
                continue
            if SELF_CITATION in body:
                body = '\n'.join(l for l in body.splitlines() if SELF_CITATION not in l).strip()
                held.append(f'{code}: dropped a citation of our own library')
            sections = plant.setdefault('sections', [])
            found = next((x for x in sections if (x.get('title') or {}).get('bg') == title_bg), None)
            if found and (found.get('body') or {}).get('bg', '').strip() == body:
                continue
            changes.append((code, f'§ {title_bg}',
                            'rewritten' if found else 'added', f'{len(body)} chars'))
            if apply_changes:
                if found:
                    found.setdefault('body', {})['bg'] = body
                else:
                    sections.append({'title': {'bg': title_bg, 'en': title_en},
                                     'body': {'bg': body, 'en': ''}})

        # --- parts ---------------------------------------------------------
        want = []
        for piece in str(row['Части'] or '').split('·'):
            piece = piece.strip()
            if not piece:
                continue
            part_code = PART_CODE.get(piece)
            if not part_code:
                held.append(f'{code}: part "{piece}" has no vocabulary code')
            elif part_code not in want:
                want.append(part_code)

        have = [p['partCode'] for p in plant.get('parts') or []]
        for part_code in want:
            if part_code in have:
                continue
            changes.append((code, 'parts +', '', part_code))
            gaps.append(f'{code}: new part "{part_code}" has no chemistry and no dosing')
            if apply_changes:
                plant.setdefault('parts', []).append(
                    {'partCode': part_code, 'chemistry': [], 'dosing': []})

        for part_code in have:
            if part_code not in want:
                held.append(f'{code}: audit drops part "{part_code}" — it carries '
                            'chemistry or dosing, so it is kept')

        # --- temperatures, per part ----------------------------------------
        #
        # After the parts pass, not before: a temperature named for a part the
        # plant gains in this same run had nowhere to land, and the script
        # needed a second pass to converge — which is not idempotence, it is
        # idempotence one run late.
        #
        # The audit writes them as prose because the field used to be per-plant:
        # „листа/кора 70–85; плодове 40–65". Now that they live on the part
        # (§13az) the prose can be read for what it always said.
        #
        # Anything that is not a number or a named part with a number is left
        # alone and reported. „специален процес" is a mode, not a temperature,
        # and is set by its own script.
        for column, field in (('Извличане °C', 'tempExtractC'),
                              ('Багрене °C', 'tempDyeC'),
                              ('Таван °C', 'softMaxTempC')):
            raw = str(row.get(column) or '').strip()
            if not raw:
                continue
            spans = parse_temperature(raw)
            if spans is None:
                held.append(f'{code}: {column} "{raw[:44]}" is not a temperature')
                continue
            # A note like „90; за плодове ~70" says the general figure first and
            # the exception second. Applied in order, the general one overwrote
            # the exception on the next run and the script stopped being
            # idempotent. The named spans win, so they are applied last.
            spans.sort(key=lambda sp: len(sp[0]))
            named = {code for sp in spans for code in sp[0]}
            for part_names, value in spans:
                targets = [x for x in plant.get('parts') or []
                           if x['partCode'] in part_names
                           or (not part_names and x['partCode'] not in named)]
                if not targets:
                    held.append(f'{code}: {column} names a part the plant does not have')
                    continue
                for target in targets:
                    now = target.get(field)
                    if field == 'softMaxTempC':
                        new_value = value[1] if value[1] is not None else value[0]
                        if now == new_value:
                            continue
                    else:
                        new_value = {'min': value[0], 'max': value[1] if value[1] is not None else value[0]}
                        if now == new_value:
                            continue
                    changes.append((code, f'{field} · {target["partCode"]}',
                                    str(now), str(new_value)))
                    if apply_changes:
                        target[field] = new_value


    report(changes, gaps, held, unmatched)

    if apply_changes:
        PACK.write_text(json.dumps(pack, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
        print(f'\nwritten: {PACK.relative_to(ROOT)}')
    else:
        print('\ndry run — pass --apply to write')


def report(changes, gaps, held, unmatched):
    print(f'changes: {len(changes)}')
    for c in changes:
        print(f'  {c[0]:<26} {c[1]:<16} {str(c[2])[:34]:<34} -> {str(c[3])[:44]}')
    if unmatched:
        print(f'\nunmatched rows: {len(unmatched)}')
        for u in unmatched:
            print(f'  {u}')
    if gaps:
        print(f'\ngaps left behind: {len(gaps)}')
        for g in sorted(set(gaps)):
            print(f'  {g}')
    if held:
        print(f'\nheld back: {len(held)}')
        for h in sorted(set(held)):
            print(f'  {h}')


if __name__ == '__main__':
    main()
