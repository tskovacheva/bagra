#!/usr/bin/env python3
"""Export every plant × part × chemistry row to a workbook for review.

The gap this addresses: of 153 chemistry entries across 57 plants, 104 carry a
`level` and 49 do not. Twenty-two plants are complete, twenty-nine partial, six
carry no chemistry at all. Partial is worse than empty — a plant showing a
strength for one compound and nothing for the next reads as though the second
were absent rather than unrecorded.

The workbook writes one row per chemistry entry, plus a row per part that has
no chemistry recorded, plus a row per plant with no parts carrying chemistry.
Only the `level` column is meant to be edited; everything else is context so
the reader can judge without opening the app.

Re-runnable. Reads only; writes the workbook and nothing else.

Usage:  python3 scripts/export-chemistry-audit.py [output.xlsx]
"""
import json
import sys
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PLANTS = 'seed/plants.json'
VOCAB = 'vocab.js'
DEFAULT_OUT = 'chemistry-audit.xlsx'

LEVELS = ['trace', 'moderate', 'high', 'dominant']
LEVEL_BG = {'trace': 'следи', 'moderate': 'умерено', 'high': 'високо', 'dominant': 'доминиращо'}


def vocab_pairs(kind):
    """Pull `code -> (bg, en)` for one vocabulary out of vocab.js.

    Read from vocab.js rather than restated here: a second copy of the
    vocabulary is a second thing to keep in step, and it would drift.
    """
    src = open(VOCAB, encoding='utf-8').read()
    out = {}
    for m in re.finditer(rf"'{kind}',\s*'([a-z_]+)',\s*'([^']*)',\s*'([^']*)'", src):
        code, bg, en = m.group(1), m.group(2), m.group(3)
        # The regex also catches the vocabulary-name list itself, whose
        # "translations" are other vocabulary names. Those have no Cyrillic.
        if re.search(r'[а-яА-Я]', bg):
            out[code] = (bg, en)
    return out


def text_of(field, lang='bg'):
    if isinstance(field, dict):
        return field.get(lang) or field.get('en') or ''
    return field or ''


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT

    plants = json.load(open(PLANTS, encoding='utf-8'))['plants']
    chem_names = vocab_pairs('chemistry_class')
    part_names = vocab_pairs('plant_part')

    rows = []
    for p in plants:
        pname = text_of(p.get('nameCommon'))
        pbot = p.get('nameBotanical') or ''
        parts = p.get('parts') or []

        if not parts:
            rows.append({
                'plant': pname, 'botanical': pbot, 'code': p['code'],
                'part': '', 'part_code': '', 'compound': '', 'compound_code': '',
                'level': '', 'status': 'растението няма части',
            })
            continue

        for part in parts:
            pc = part.get('partCode', '')
            part_bg = part_names.get(pc, (pc, pc))[0]
            chem = part.get('chemistry') or []

            if not chem:
                rows.append({
                    'plant': pname, 'botanical': pbot, 'code': p['code'],
                    'part': part_bg, 'part_code': pc,
                    'compound': '', 'compound_code': '',
                    'level': '', 'status': 'няма записана химия',
                })
                continue

            for c in chem:
                cc = c.get('classCode', '')
                lvl = c.get('level') or ''
                rows.append({
                    'plant': pname, 'botanical': pbot, 'code': p['code'],
                    'part': part_bg, 'part_code': pc,
                    'compound': chem_names.get(cc, (cc, cc))[0], 'compound_code': cc,
                    'level': lvl,
                    'status': '' if lvl else 'ЗА ПОПЪЛВАНЕ',
                })

    # Sort so the work clusters, and so the two KINDS of work stay apart. They
    # are different jobs: filling a level is a judgement about a compound
    # already known to be present, whereas a part with no chemistry at all is a
    # question about whether anything belongs there. Interleaved, the second
    # kind reads as more of the first and gets guessed at.
    #
    # Sorted in Python because the workbook must not depend on a spilling
    # function to be readable.
    order = {'ЗА ПОПЪЛВАНЕ': 0, 'няма записана химия': 1, 'растението няма части': 2}
    rows.sort(key=lambda r: (order.get(r['status'], 3), r['plant'], r['part_code'], r['compound']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Химия'

    headers = ['Растение', 'Латинско име', 'Код', 'Част', 'Вещество',
               'Степен', 'Състояние', 'Бележка']
    widths = [24, 26, 22, 18, 22, 16, 20, 34]

    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2C3B57')   # indigo, per palette
    body_font = Font(name='Arial', size=10)
    edit_fill = PatternFill('solid', fgColor='FFF3CD')     # cells to fill in
    flag_font = Font(name='Arial', size=10, bold=True, color='A03D3B')  # madder
    thin = Side(style='thin', color='DED8CA')
    border = Border(bottom=thin)

    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for n, r in enumerate(rows, start=2):
        values = [r['plant'], r['botanical'], r['code'], r['part'],
                  r['compound'], r['level'], r['status'], '']
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=n, column=i, value=v)
            cell.font = body_font
            cell.border = border
        if r['status'] == 'ЗА ПОПЪЛВАНЕ':
            ws.cell(row=n, column=6).fill = edit_fill
            ws.cell(row=n, column=7).font = flag_font
        elif r['status']:
            ws.cell(row=n, column=7).font = flag_font

    # A dropdown rather than free text: the four levels are a controlled
    # vocabulary in vocab.js, and a typo here becomes a silent unknown value
    # on import.
    dv = DataValidation(
        type='list',
        formula1='"' + ','.join(LEVELS) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = 'Само: trace, moderate, high, dominant'
    dv.errorTitle = 'Невалидна степен'
    ws.add_data_validation(dv)
    dv.add(f'F2:F{len(rows) + 1}')

    # ---------------------------------------------------------------- legend
    ls = wb.create_sheet('Как се попълва')
    ls.column_dimensions['A'].width = 22
    ls.column_dimensions['B'].width = 82

    def line(row, a, b, bold=False):
        ca = ls.cell(row=row, column=1, value=a)
        cb = ls.cell(row=row, column=2, value=b)
        ca.font = Font(name='Arial', size=10, bold=True)
        cb.font = Font(name='Arial', size=10, bold=bold)
        cb.alignment = Alignment(wrap_text=True, vertical='top')

    total = len(rows)
    to_fill = sum(1 for r in rows if r['status'] == 'ЗА ПОПЪЛВАНЕ')
    no_chem = sum(1 for r in rows if r['status'] == 'няма записана химия')

    line(1, 'Какво е това', 'Всички химични записи в библиотеката, по растение и част.')
    line(3, 'Какво се пипа', 'САМО колона «Степен» (жълтите клетки). Останалото е контекст.')
    line(4, 'Позволени стойности', 'trace (следи) · moderate (умерено) · high (високо) · dominant (доминиращо)')
    line(5, '', 'Клетката има падащ списък — не се пише свободен текст.')
    line(7, 'Редове общо', str(total))
    line(8, 'За попълване', f'{to_fill} — жълти, подредени най-отгоре')
    line(9, 'Части без химия', f'{no_chem} — трябва решение: има ли изобщо какво да се запише')
    line(11, 'Ако не се знае', 'Остави празно. Празно означава «неизвестно» и си остава честно.')
    line(12, '', 'Измислена степен влиза в справочника като факт — по-лошо от липсваща.')
    line(14, 'Колона «Бележка»', 'Свободна. За източник, съмнение, или «това вещество не е тук».')
    line(16, 'Обратно в приложението', 'Върни файла — има скрипт, който го слива обратно в seed/plants.json.')

    for r in (3, 4, 5, 11, 12, 14, 16):
        ls.row_dimensions[r].height = 15

    wb.save(out_path)
    print(f'{out_path}: {total} реда, {to_fill} за попълване, {no_chem} части без химия')


if __name__ == '__main__':
    main()
